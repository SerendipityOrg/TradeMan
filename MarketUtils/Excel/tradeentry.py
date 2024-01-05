import os,sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment,NamedStyle

import strategy_calc as sc
import firebase_admin
from firebase_admin import credentials, storage
from telethon.sync import TelegramClient
from dotenv import load_dotenv


DIR = os.getcwd()
sys.path.append(DIR)

import MarketUtils.general_calc as general_calc

ENV_PATH = os.path.join(DIR, '.env')
load_dotenv(ENV_PATH)

api_id = os.getenv('telethon_api_id')
api_hash = os.getenv('telethon_api_hash')
excel_dir = os.getenv('onedrive_excel_folder')

broker_filepath = os.path.join(DIR,"MarketUtils", "broker.json")
userprofile_dir = os.path.join(DIR, "UserProfile","OrdersJson")
active_users_filepath = os.path.join(DIR,"MarketUtils", "active_users.json")
credentials_filepath = os.path.join(DIR,"MarketUtils","Excel","credentials.json")


class TradingStrategy:
    def __init__(self, name, process_func):
        self.name = name
        self.process_func = process_func
    
    def process_data(self, user_data, broker,username):
        if self.name in user_data["today_orders"]:
            data = self.process_func(broker, user_data["today_orders"][self.name],username)
            df = pd.DataFrame(data)
            if 'pnl' in df.columns:
                PnL = round(df['pnl'].sum(), 1)
                Tax = round(df['tax'].sum(), 1)
            else:
                print("PnL column not found in DataFrame")
                PnL = 0
                Tax = 0
            return df, PnL, Tax
        return pd.DataFrame(), 0, 0

def load_existing_excel(excel_path):
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    try:
        book = load_workbook(excel_path)
        return {sheet_name: pd.read_excel(excel_path, sheet_name=sheet_name) for sheet_name in book.sheetnames}
    except Exception as e:
        print(f"An error occurred while loading the Excel file: {excel_path}")
        print("Error:", e)
        return {}

def update_excel_data(all_dfs, df, strategy_name, unique_id_column='trade_id'):
    if not df.empty:
        # Check if the strategy's DataFrame exists in all_dfs
        if strategy_name in all_dfs:
            strategy_df = all_dfs[strategy_name]
            for i, row in df.iterrows():
                unique_id = row[unique_id_column]
                # Find the index in the existing DataFrame
                index = strategy_df[strategy_df[unique_id_column] == unique_id].index
                if not index.empty:
                    # Update the existing row (ensuring columns match)
                    for col in strategy_df.columns:
                        strategy_df.at[index[0], col] = row[col] if col in row else strategy_df.at[index[0], col]
                else:
                    # Append new row if the unique id does not exist
                    strategy_df = pd.concat([strategy_df, pd.DataFrame([row])], ignore_index=True)
            all_dfs[strategy_name] = strategy_df
        else:
            # If the strategy's DataFrame does not exist, simply add it
            all_dfs[strategy_name] = df

def save_all_sheets_to_excel(all_dfs, excel_path):
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        if 'number_style' not in writer.book.named_styles:
            number_style = NamedStyle(name='number_style', number_format='0.00')
            writer.book.add_named_style(number_style)
        
        center_alignment = Alignment(horizontal='center')

        for sheet_name, df in all_dfs.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]

            # Apply number formatting and center alignment to specified columns
            rounded_columns = ['entry_price', 'exit_price', 'hedge_entry_price', 'hedge_exit_price', 'trade_points', 'pnl', 'tax', 'net_pnl']

            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    # Apply center alignment to all cells
                    cell.alignment = center_alignment

                    # Apply number formatting to specified columns
                    if cell.col_idx - 1 < len(df.columns) and df.columns[cell.col_idx - 1] in rounded_columns:
                        cell.number_format = number_style.number_format



def build_message(user, strategy_results, gross_pnl, tax, current_capital, expected_capital):
    message_parts = [f"Hello {user},We hope you're enjoying a wonderful day.\n Here are your PNLs for today:\n"]
    
    for strategy_name, pnl in strategy_results.items():
        if pnl != 0:
            message_parts.append(f"{strategy_name}: {sc.custom_format(pnl)}")
    
    message_parts.extend([
        f"\n**Gross PnL: {sc.custom_format(gross_pnl)}**",
        f"**Expected Tax: {sc.custom_format(tax)}**",
        f"**Current Capital: {sc.custom_format(current_capital)}**",
        f"**Expected Morning Balance : {sc.custom_format(expected_capital)}**",
        "\nBest Regards,\nSerendipity Trading Firm"
    ])
    
    return message_parts

# Define a dictionary that maps strategy names to their processing functions
strategy_config = {
    "MPWizard": sc.process_mpwizard_trades,
    "AmiPy": sc.process_amipy_trades,
    "OvernightFutures": sc.process_overnight_futures_trades,
    "ExpiryTrader": sc.process_expiry_trades,
    "Extra": sc.process_extra_trades
    # Add new strategies here as needed
}

cred = credentials.Certificate(credentials_filepath)
firebase_admin.initialize_app(cred, {
    'databaseURL': 'https://trading-app-caf8e-default-rtdb.firebaseio.com'
})

def save_to_firebase(user, excel_path):
    # Correct bucket name
    bucket = storage.bucket(name='trading-app-caf8e.appspot.com')
    blob = bucket.blob(f'{user}.xlsx')
    with open(excel_path, 'rb') as my_file:
        blob.upload_from_file(my_file)
    print(f"Excel file for {user} has been uploaded to Firebase.")

def main():
    data = general_calc.read_json_file(active_users_filepath)
    
    # Initialize TradingStrategy objects based on the configuration
    strategies = [TradingStrategy(name, func) for name, func in strategy_config.items()]

    for user in data:
        username = user['account_name']
        user_data = general_calc.read_json_file(os.path.join(userprofile_dir, f"{user['account_name']}.json"))
        broker_json = general_calc.read_json_file(broker_filepath)
        broker = user["broker"]

        strategy_results = {}
        gross_pnl = 0
        total_tax = 0
        
        excel_path = os.path.join(excel_dir, f"{user['account_name']}.xlsx")
        all_dfs = load_existing_excel(excel_path)

        for strategy in strategies:
                df, pnl, tax = strategy.process_data(user_data, broker,username)
                strategy_results[strategy.name] = pnl
                gross_pnl += pnl
                total_tax += tax
                update_excel_data(all_dfs, df, strategy.name)
                
        net_pnl = gross_pnl - total_tax

        for account in broker_json:
            if account["account_name"] == user["account_name"]:
                current_capital = account["expected_morning_balance"]

        expected_capital = current_capital + net_pnl if net_pnl > 0 else current_capital - abs(net_pnl)

        message_parts = build_message(user['account_name'], strategy_results, gross_pnl, total_tax, current_capital, expected_capital)
        message = "\n".join(message_parts).replace('\u20b9', '₹')
        print(message)

        save_all_sheets_to_excel(all_dfs, excel_path)
        save_to_firebase(username,excel_path)

if __name__ == "__main__":
    main()
