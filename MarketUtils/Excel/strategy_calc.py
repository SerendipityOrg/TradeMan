import pandas as pd
import os,sys
from babel.numbers import format_currency
from openpyxl import load_workbook
from dotenv import load_dotenv

DIR = os.getcwd()
sys.path.append(DIR)
import MarketUtils.Calculations.taxcalculation as tc

ENV_PATH = os.path.join(DIR, '.env')
load_dotenv(ENV_PATH)
excel_dir = os.getenv('onedrive_excel_folder')

def custom_format(amount):
    formatted = format_currency(amount, 'INR', locale='en_IN')
    return formatted.replace('₹', '₹')

def load_existing_excel(excel_path):
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    try:
        book = load_workbook(excel_path)
        return {sheet_name: pd.read_excel(excel_path, sheet_name=sheet_name) for sheet_name in book.sheetnames}
    except Exception as e:
        print(f"An error occurred while loading the Excel file: {excel_path}")
        print("Error:", e)
        return {}

def process_mpwizard_trades(broker,mpwizard_trades,username=None):
    if not mpwizard_trades:
        print("No MPWizard trades found.")
        return []

    result = []

    for i in range(len(mpwizard_trades["BUY"])):
        buy_trade = mpwizard_trades["BUY"][i]
        # Find a matching sell_trade by trade_id or trading_symbol
        matching_sell_trades = [trade for trade in mpwizard_trades["SELL"]
                                if trade["trade_id"] == buy_trade["trade_id"] or
                                trade["trading_symbol"] == buy_trade["trading_symbol"]]
        
        if not matching_sell_trades:
            print(f"No matching SELL trade for BUY trade with trade_id: {buy_trade['trade_id']}")
            continue
        
        sell_trade = matching_sell_trades[0] # Take the first matching sell trade

        if broker == "zerodha":
            charges = tc.zerodha_taxes(
                buy_trade["qty"], buy_trade["avg_price"], sell_trade["avg_price"], 1)
        elif broker == "aliceblue":
            charges = tc.aliceblue_taxes(
                buy_trade["qty"], float(buy_trade["avg_price"]), float(sell_trade["avg_price"]), 1)
        
        trade_points = float(sell_trade["avg_price"]) - float(buy_trade["avg_price"])
        pnl = trade_points * int(buy_trade["qty"])
        net_pnl = pnl - charges
        signal = "Long"

        entry_time = pd.to_datetime(buy_trade["time"], format='%d/%m/%Y %H:%M:%S').round('min')
        exit_time = pd.to_datetime(sell_trade["time"], format='%d/%m/%Y %H:%M:%S').round('min')

        trade_data = {
            "trade_id": buy_trade["trade_id"],
            "trading_symbol": buy_trade["trading_symbol"],
            "signal": signal,
            "entry_time": entry_time.strftime('%Y-%m-%d %H:%M:%S'),
            "exit_time": exit_time.strftime('%Y-%m-%d %H:%M:%S'),
            "entry_price": round(buy_trade["avg_price"], 2),
            "exit_price": round(sell_trade["avg_price"], 2),
            "hedge_entry_price": 0,  # Assuming no hedge for this example
            "hedge_exit_price": 0,   # Assuming no hedge for this example
            "trade_points": round(trade_points, 2),
            "qty": buy_trade["qty"],
            "pnl": round(pnl, 2),
            "tax": round(charges, 2),
            "net_pnl": round(net_pnl, 2)
        }
        result.append(trade_data)

    return result

def process_amipy_trades(broker,amipy_trades,username=None):
    amipy_data_short = []
    amipy_data_long = []
    if "ShortSignal" in amipy_trades:
        # Process the AmiPy ShortSignal and ShortCoverSignal trades
        amipy_data_short = process_short_trades(broker,amipy_trades["ShortSignal"],
                                                amipy_trades["ShortCoverSignal"],"Short")
    if "LongSignal" in amipy_trades:
        # Process the AmiPy LongSignal and LongCoverSignal trades
        amipy_data_long = process_long_trades(broker,amipy_trades["LongSignal"],
                                                amipy_trades["LongCoverSignal"],"Long")
    if amipy_data_short is not None:
        return amipy_data_short
    elif amipy_data_long is not None:
        return amipy_data_long
    
def process_short_trades(broker: str, short_signals: list[dict], short_cover_signals: list[dict], signal: str) -> None:
    if len(short_signals) != len(short_cover_signals):
        print("Mismatch in the number of ShortSignal and ShortCoverSignal trades.")
        return []

    result = []
    for i in range(0, len(short_signals), 4):  # Process each group of 4 trades together
        short_signal_group = short_signals[i:i+4]
        short_cover_signal_group = short_cover_signals[i:i+4]

        hedge_entry = sum(float(
            trade["avg_price"]) for trade in short_signal_group if trade["trade_type"] == "HedgeOrder")
        hedge_exit = sum(float(
            trade["avg_price"]) for trade in short_cover_signal_group if trade["trade_type"] == "HedgeOrder")

        entry_price = sum(float(
            trade["avg_price"]) for trade in short_signal_group if trade["trade_type"] == "ShortSignal")
        exit_price = sum(float(
            trade["avg_price"]) for trade in short_cover_signal_group if trade["trade_type"] == "ShortCoverSignal")
        hedge_price = hedge_exit - hedge_entry
        trade_points = (entry_price - exit_price) + hedge_price
        

        if broker == "zerodha":
            charges = tc.zerodha_taxes(
                short_signal_group[0]["qty"], entry_price, exit_price, 2)
            hedge_charges = tc.zerodha_taxes(
                short_signal_group[0]["qty"], hedge_entry, hedge_exit, 2)
        elif broker == "aliceblue":
            charges = tc.aliceblue_taxes(
                short_signal_group[0]["qty"], entry_price, exit_price, 2)
            hedge_charges = tc.aliceblue_taxes(
                short_signal_group[0]["qty"], hedge_entry, hedge_exit, 2)
        charges = charges + hedge_charges

        entry_time = pd.to_datetime(short_signal_group[0]["time"], format='%d/%m/%Y %H:%M:%S').round('min')
        exit_time = pd.to_datetime(short_cover_signal_group[0]["time"], format='%d/%m/%Y %H:%M:%S').round('min')

        trade_data = {
            "trade_id": short_signal_group[0]["trade_id"],
            "trading_symbol": short_signal_group[0]["trading_symbol"],
            "signal": signal,
            "entry_time": entry_time.strftime('%Y-%m-%d %H:%M:%S'),
            "exit_time": exit_time.strftime('%Y-%m-%d %H:%M:%S'),
            "entry_price": round(entry_price,2),
            "exit_price": round(exit_price,2),
            "hedge_entry_price": round(hedge_entry,2),
            "hedge_exit_price": round(hedge_exit,2),
            "trade_points": round(trade_points,2),
            "qty": short_signal_group[0]["qty"],
            "pnl": round(trade_points * int(short_signal_group[0]["qty"]),2),
            "tax": round(charges,2),
            "net_pnl" : round((trade_points * int(short_signal_group[0]["qty"]) - charges),2)
        }           
        result.append(trade_data)
    return result

def process_long_trades(broker,long_signals, long_cover_signals,signal):
    if len(long_signals) != len(long_cover_signals):
        print("Mismatch in the number of LongSignal and LongCoverSignal trades.")
        return []

    result = []
    for i in range(0, len(long_signals), 2):  # Process each pair of trades together
        long_signal_group = long_signals[i:i+2]
        long_cover_signal_group = long_cover_signals[i:i+2]

        entry_price = sum(float(trade["avg_price"])
                          for trade in long_signal_group)
        exit_price = sum(float(trade["avg_price"])
                         for trade in long_cover_signal_group)
        trade_points = entry_price - exit_price

        if broker == "zerodha":
            charges = tc.zerodha_taxes(
                long_signal_group[0]["qty"], entry_price, exit_price, 2)
        elif broker == "aliceblue":
            charges = tc.aliceblue_taxes(
                long_signal_group[0]["qty"], entry_price, exit_price, 2)

        entry_time = pd.to_datetime(long_signal_group[0]["time"], format='%d/%m/%Y %H:%M:%S').round('min')
        exit_time = pd.to_datetime(long_cover_signal_group[0]["time"], format='%d/%m/%Y %H:%M:%S').round('min')

        trade_data = {
            "trade_id": long_signal_group[0]["trade_id"],
            "trading_symbol": long_signal_group[0]["trading_symbol"],
            "signal": signal,
            "entry_time": entry_time.strftime('%Y-%m-%d %H:%M:%S'),
            "exit_time": exit_time.strftime('%Y-%m-%d %H:%M:%S'),
            "entry_price": round(entry_price,2),
            "entry_price": round(exit_price,2),
            "hedge_entry_price": 0,
            "hedge_exit_price": 0,
            "trade_points": round(trade_points,2),
            "qty": long_signal_group[0]["qty"],
            "pnl": round(trade_points * int(long_signal_group[0]["qty"]),2),
            "tax": round(charges,2),
            "net_pnl" : round((trade_points * int(long_signal_group[0]["qty"]) - charges),2)
        }
        result.append(trade_data)
    return result

def process_expiry_trades(broker, expiry_trades,username=None):
    def calculate_avg_price(trades):
        total_price = sum(trade["avg_price"] * trade["qty"] for trade in trades)
        total_qty = sum(trade["qty"] for trade in trades)
        return total_price / total_qty if total_qty > 0 else 0

    if not expiry_trades:
        print("No ExpiryTrades trades found.")
        return []

    result = []
    trade_ids = set(trade["trade_id"] for trade in expiry_trades["Entry"])

    for trade_id in trade_ids:
        entry_trades = [trade for trade in expiry_trades["Entry"] if trade["trade_id"] == trade_id]
        exit_trades = [trade for trade in expiry_trades["Exit"] if trade["trade_id"] == trade_id]

        main_entry_trades = [trade for trade in entry_trades if trade["trade_type"] == "MainOrder"]
        hedge_entry_trades = [trade for trade in entry_trades if trade["trade_type"] == "HedgeOrder"]
        main_exit_trades = [trade for trade in exit_trades if trade["trade_type"] == "MainOrder"]
        hedge_exit_trades = [trade for trade in exit_trades if trade["trade_type"] == "HedgeOrder"]

        main_entry_avg_price = calculate_avg_price(main_entry_trades)
        hedge_entry_avg_price = calculate_avg_price(hedge_entry_trades)
        main_exit_avg_price = calculate_avg_price(main_exit_trades)
        hedge_exit_avg_price = calculate_avg_price(hedge_exit_trades)

        # Number of orders (assuming 2 trades per order)
        no_of_orders = len(entry_trades) // 2

        if broker == "zerodha":
            charges = tc.zerodha_taxes(main_entry_trades[0]["qty"], main_entry_avg_price, main_exit_avg_price, no_of_orders)
        elif broker == "aliceblue":
            charges = tc.aliceblue_taxes(main_entry_trades[0]["qty"], main_entry_avg_price, main_exit_avg_price, no_of_orders)
        else:
            charges = 0  # No charges if broker is not recognized

        main_trade_points = main_entry_avg_price - main_exit_avg_price
        hedge_trade_points = hedge_exit_avg_price - hedge_entry_avg_price if hedge_entry_trades else 0
        trade_points = main_trade_points + hedge_trade_points
        pnl = trade_points * main_entry_trades[0]["qty"]
        net_pnl = pnl - charges


        # Parse dates with the correct format
        entry_time = pd.to_datetime(main_entry_trades[0]["time"], format='%d/%m/%Y %H:%M:%S').round('min')
        exit_time = pd.to_datetime(main_exit_trades[0]["time"], format='%d/%m/%Y %H:%M:%S').round('min')

        trade_data = {
            "trade_id": trade_id,
            "trading_symbol": main_entry_trades[0]["trading_symbol"],
            "signal": "Short",
            "entry_time": entry_time.strftime('%Y-%m-%d %H:%M:%S'),
            "exit_time": exit_time.strftime('%Y-%m-%d %H:%M:%S'),
            "entry_price": round(main_entry_avg_price, 2),
            "exit_price": round(main_exit_avg_price, 2),
            "hedge_entry_price": round(hedge_entry_avg_price, 2) if hedge_entry_trades else 0,
            "hedge_exit_price": round(hedge_exit_avg_price, 2) if hedge_exit_trades else 0,
            "trade_points": round(trade_points, 2),
            "qty": main_entry_trades[0]["qty"],
            "pnl": round(pnl, 2),
            "tax": round(charges, 2),
            "net_pnl": round(net_pnl, 2)
        }

        result.append(trade_data)

    return result

def process_morning_trades(broker,morning_trades,username=None):
    excel_path = os.path.join(excel_dir, f"{username}.xlsx")
    all_dfs = load_existing_excel(excel_path)
    trade_df = all_dfs.get("OvernightFutures", pd.DataFrame())
    trade_id = morning_trades[0]['trade_id'].split("_")[0]
    trade_index = trade_df.index[trade_df['trade_id'] == trade_id].tolist()

    if trade_index:
        row_index = trade_index[0]
        trade_data = trade_df.loc[row_index]
        #delete the row_index from the excel
        trade_df.drop(row_index,inplace=True)

        for trade in morning_trades:
            if trade['option_type'] != "FUT":
                option_exit_price = trade['avg_price']
            else:
                future_exit_price = trade['avg_price']
        
        qty = trade['qty']

        if broker == "zerodha":
            future_tax = tc.zerodha_futures_taxes(qty, trade_data['entry_price'], future_exit_price, 1)
            option_tax = tc.zerodha_taxes(qty, trade_data["hedge_entry_price"],option_exit_price, 1)
        elif broker == "aliceblue":
            future_tax = tc.aliceblue_futures_taxes(qty, trade_data["entry_price"], future_exit_price, 1)
            option_tax = tc.aliceblue_taxes(qty, trade_data["hedge_entry_price"], option_exit_price, 1)

        # Calculating trade points based on direction
        if trade_data["signal"] == "Long":
            trade_points = (future_exit_price - trade_data["entry_price"]) + (option_exit_price - trade_data["hedge_entry_price"])
        elif trade_data["signal"] == "Short":
            trade_points = (trade_data["entry_price"] - future_exit_price) + (option_exit_price - trade_data["hedge_entry_price"])
        pnl = trade_points * qty
        total_tax = future_tax + option_tax
        net_pnl = pnl - total_tax

        morning_trade_data = {
                    "trade_id": trade_data["trade_id"],
                    "trading_symbol": trade_data["trading_symbol"],
                    "signal": trade_data["signal"],
                    "entry_time": trade_data["entry_time"],
                    "exit_time": pd.to_datetime(trade["time"], format='%d/%m/%Y %H:%M:%S').round('min'),
                    "entry_price": trade_data["entry_price"],
                    "exit_price": future_exit_price,
                    "hedge_entry_price": trade_data["hedge_entry_price"],
                    "hedge_exit_price": option_exit_price,
                    "trade_points": trade_points,
                    "qty": trade_data["qty"],
                    "pnl": pnl,
                    "tax": total_tax,
                    "net_pnl": net_pnl
                }

    return morning_trade_data

def process_afternoon_trades(broker,afternoon_trades,username=None):
    signal = "Long" if afternoon_trades[0]['direction'] == "BULLISH" else "Short"
    for trade in afternoon_trades:
        if trade['option_type'] != "FUT":
            option_entry_price = trade['avg_price']

        else:
            future_entry_price = trade['avg_price']
    
    afternoon_trade_data = {
                    "trade_id": afternoon_trades[0]['trade_id'],
                    "trading_symbol": afternoon_trades[0]['trading_symbol'],
                    "signal": signal,
                    "entry_time": pd.to_datetime(trade['time'], format='%d/%m/%Y %H:%M:%S').round('min'),
                    "entry_price": future_entry_price,
                    "hedge_entry_price": option_entry_price,
                    "qty": afternoon_trades[0]['qty'],
                    "pnl": 0,
                    "tax": 0
                }
    return afternoon_trade_data

def process_overnight_futures_trades(broker,trades,username=None):
    if not trades:
        print("No OvernightFutures trades found.")
        return []
    
    result = []

    if trades['Morning']:
        morning_trade_data = process_morning_trades(broker,trades['Morning'],username)
        result.append(morning_trade_data)
    
    if trades['Afternoon']:
        afternoon_trade_data = process_afternoon_trades(broker,trades['Afternoon'],username)
        result.append(afternoon_trade_data)
    
    return result

def process_extra_trades(broker,extra_trades,username=None):
    if not extra_trades:
        print("No extra trades found.")
        return []
    
    def find_matching_entry_trade(processed_trades, exit_trade):
        """
        Find a matching entry trade for the given exit trade.
        """
        for entry_trade in processed_trades:
            if entry_trade['trade_id'] == exit_trade['trade_id']:
                return entry_trade
        return None
    
    processed_trades = []

    def process_entry_trade(trade, trade_type):
        trade_id = trade['trade_id'].split("_")[0]
        entry_trade = {
            "trade_id": trade_id,
            "trading_symbol": trade['trading_symbol'],
            "signal": trade_type,
            "entry_time": pd.to_datetime(trade["time"], format='%d/%m/%Y %H:%M:%S').round('min'),
            "entry_price": trade['avg_price'],
            "hedge_entry_price": 0.00,
            "qty": trade['qty']
        }
        return entry_trade
    
    def process_exit_trade(trade, matching_trade):
        if matching_trade:
            trade_points = trade['avg_price'] - matching_trade['entry_price']
            pnl = trade_points * trade['qty']

            if broker == "zerodha":
                charges = tc.zerodha_taxes(trade['qty'], matching_trade['entry_price'], trade['avg_price'], 1)
            elif broker == "aliceblue":
                charges = tc.aliceblue_taxes(trade['qty'], matching_trade['entry_price'], trade['avg_price'], 1)
            net_pnl = pnl - charges
            trade_id = trade['trade_id'].split("_")[0]
            exit_trade = {
                "trade_id": trade_id,
                "trading_symbol": trade['trading_symbol'],
                "signal": matching_trade['signal'],
                "entry_time": matching_trade['entry_time'],
                "exit_time": pd.to_datetime(trade["time"], format='%d/%m/%Y %H:%M:%S').round('min'),
                "entry_price": matching_trade['entry_price'],
                "exit_price": trade['avg_price'],
                "hedge_entry_price": 0.00,
                "hedge_exit_price": 0.00,
                "trade_points": trade_points,
                "qty": trade['qty'],
                "pnl": pnl,
                "tax": charges,
                "net_pnl": net_pnl
            }
            return exit_trade
    

    def fetch_trade_details_from_excel(trade, cover_type,username):
        excel_path = os.path.join(DIR, f"UserProfile/Excel/{username}.xlsx")
        all_dfs = load_existing_excel(excel_path)
        trade_df = all_dfs.get("Extra", pd.DataFrame())
        trade_id = trade['trade_id'].split("_")[0]
        trade_index = trade_df.index[trade_df['trade_id'] == trade_id].tolist()

        if trade_index:
            row_index = trade_index[0]
            trade_data = trade_df.loc[row_index]
            
            if trade_data is not None:
                if cover_type == "LongCover":
                    trade_points = trade['avg_price'] - trade_data['entry_price']
                elif cover_type == "ShortCover":
                    trade_points = trade_data['entry_price'] - trade['avg_price']
                pnl = trade_points * trade['qty']
                if broker == "zerodha":
                    charges = tc.zerodha_taxes(trade['qty'], trade_data['entry_price'], trade['avg_price'], 1)
                elif broker == "aliceblue":
                    charges = tc.aliceblue_taxes(trade['qty'], trade_data['entry_price'], trade['avg_price'], 1)
                net_pnl = pnl - charges

                extra_trade_data = {
                    "trade_id": trade_data["trade_id"],
                    "trading_symbol": trade_data["trading_symbol"],
                    "signal": trade_data["signal"],
                    "entry_time": trade_data["entry_time"],
                    "exit_time": pd.to_datetime(trade["time"], format='%d/%m/%Y %H:%M:%S').round('min'),
                    "entry_price": trade_data["entry_price"],
                    "exit_price": trade['avg_price'],
                    "hedge_entry_price": 0.00,
                    "hedge_exit_price": 0.00,
                    "trade_points": trade_points,
                    "qty": trade_data["qty"],
                    "pnl": pnl,
                    "tax": charges,
                    "net_pnl": net_pnl
                }
                return extra_trade_data
        else:
            print(f"No matching {cover_type} trade for trade with trade_id: {trade['trade_id']}")
            return None
        

    # Assuming extra_trades is a dictionary with keys: 'Long', 'Short', 'LongCover', 'ShortCover'
    for trade_type in ['Long', 'Short']:
        for trade in extra_trades[trade_type]:
            processed_trade = process_entry_trade(trade, trade_type)
            processed_trades.append(processed_trade)

    for cover_type in ['LongCover', 'ShortCover']:
        for trade in extra_trades[cover_type]:
            matching_trade = find_matching_entry_trade(processed_trades, trade)
            if matching_trade:
                processed_trade = process_exit_trade(trade, matching_trade)
            else:
                processed_trade = fetch_trade_details_from_excel(trade, cover_type,username)
            processed_trades.append(processed_trade)
    
    return processed_trades