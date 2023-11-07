import os, sys
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from babel.numbers import format_currency
from telethon.sync import TelegramClient
from dotenv import load_dotenv


# Setting up directory paths
DIR = os.getcwd()
sys.path.append(DIR)

import MarketUtils.general_calc as general_calc
import Streamlitapp.formats as custom_format
import dtdautomation as dtd

script_dir = os.path.dirname(os.path.realpath(__file__))
broker_filepath = os.path.join(DIR, "MarketUtils", "broker.json")
excel_dir = os.path.join(DIR, "UserProfile", "excel")
ENV_PATH = os.path.join(DIR,'Streamlitapp', '.env')

# Loading environment variables
load_dotenv(ENV_PATH)
api_id = os.getenv('telethon_api_id')
api_hash = os.getenv('telethon_api_hash')


# Function to load an existing Excel file and return its data as a dictionary of pandas DataFrames
def load_existing_excel(excel_path):
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    try:
        book = load_workbook(excel_path)
        return {sheet_name: pd.read_excel(excel_path, sheet_name=sheet_name) for sheet_name in book.sheetnames}
    except Exception as e:
        print(f"An error occurred while loading the Excel file: {excel_path}")
        print("Error:", e)
        return {}


# Function to send a message via Telegram
def send_telegram_message(phone_number, message):
    session_filepath = os.path.join(script_dir, "..", '..', '..', "+918618221715.session")
    with TelegramClient(session_filepath, api_id, api_hash) as client:
        client.send_message(phone_number, message, parse_mode='md')


# The main function which processes user data and sends PNL messages
def main():
    broker_data = general_calc.read_json_file(broker_filepath)
    active_users = [user for user in broker_data if "Active" in user['account_type']]
    today = datetime.now().strftime('%Y-%m-%d')

    for user in active_users:
        excel_path = os.path.join(excel_dir, f"{user['account_name']}.xlsx")
        all_dfs = load_existing_excel(excel_path)

        strategy_results = {}
        gross_pnl = 0
        total_tax = 0

        for strategy, df in all_dfs.items():
            if 'entry_time' in df.columns and 'exit_time' in df.columns:
                df_today = df[(df['entry_time'] == today) | (df['exit_time'] == today)]
                if not df_today.empty:
                    pnl = df_today['pnl'].sum()
                    tax = df_today['tax'].sum()

                    strategy_results[strategy] = pnl
                    gross_pnl += pnl
                    total_tax += tax

        net_pnl = gross_pnl - total_tax
        current_capital = next(account["expected_morning_balance"] for account in broker_data if account["account_name"] == user["account_name"])
        expected_capital = current_capital + net_pnl if net_pnl > 0 else current_capital - abs(net_pnl)
        
        message = build_message(user['account_name'], strategy_results, gross_pnl, total_tax, current_capital, expected_capital)
        print(message)
        
        # Sending the message via Telegram
        send_telegram_message(user['phone_number'], message)


# Function to update JSON data with new PNL, capital values, etc.
def update_json_data(data, broker, user, net_pnl, current_capital, expected_capital, broker_filepath):
    for username in data:
        if user["account_name"] == username["account_name"]:
            if "Active" in user['account_type']:
                user_details = username
                user_details["current_capital"] = round(current_capital, 2)
                user_details["yesterday_PnL"] = net_pnl
                user_details["expected_morning_balance"] = round(expected_capital, 2)
    general_calc.write_json_file(broker_filepath, data)


# Function to construct a PNL message for the user
def build_message(user, strategy_results, gross_pnl, tax, current_capital, expected_capital,total_tax):
    message_parts = [
        f"Hello {user}, We hope you're enjoying a wonderful day.\n Here are your PNLs for today:\n"
    ]

    for strategy_name, pnl in strategy_results.items():
        if pnl != 0:
            message_parts.append(f"{strategy_name}: {custom_format(pnl)}")

    message_parts.extend([
        f"\n**Gross PnL: {custom_format(gross_pnl)}**",
        f"**Expected Tax: {custom_format(tax)}**",
        f"**Current Capital: {custom_format(current_capital)}**",
        f"**Expected Morning Balance : {custom_format(expected_capital)}**",
        "\nBest Regards,\nSerendipity Trading Firm"
    ])

    return "\n".join(message_parts).replace('\u20b9', '₹')


# Execute the main function when script is run
if __name__ == "__main__":
    main()
