import json
import os

from openpyxl import load_workbook


def find_first_empty_row(sheet):
    """
    Find the first empty row in a given worksheet.
    """
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if all(cell is None for cell in row):
            return i
    return sheet.max_row + 1

def is_order_present(sheet, tradingsymbol, timestamp):
    """
    Check if an order with a given tradingsymbol and timestamp is already present in the worksheet.
    """
    for row in sheet.iter_rows(values_only=True):
        if row and row[1] == tradingsymbol and row[2] == timestamp:
            return True
    return False

def get_user_list(data):
    """
    Populate user_list with accounts from each broker and return the populated list.
    """
    user_list = [(broker, account) for broker, broker_data in data.items() for account in broker_data.get('accounts_to_trade', [])]
    return user_list
