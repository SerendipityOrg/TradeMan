import json
import os
from pprint import pprint

from openpyxl import load_workbook
from pya3 import *

from .holdings_utils import (find_first_empty_row, get_user_list,
                             is_order_present)

# File paths
script_dir = os.path.dirname(os.path.realpath(__file__))
broker_filepath = os.path.join(script_dir, "broker.json")

# Load the broker data from the JSON file
with open(broker_filepath) as file:
    data = json.load(file)

def fetch_aliceblue_orders(username, api_key):
    """
    Fetch orders from Alice Blue using the smartapi-python library.
    """
    alice = Aliceblue(username, api_key)
    session_id = alice.get_session_id()
    holdings = alice.get_holding_positions()
    pprint(holdings)
    orders = alice.get_daywise_positions()
    
    # If the data type of orders is not a list, return an empty list
    if not isinstance(orders, list):
        print("Unexpected format for orders. Expected a list but got:", type(orders))
        return []
    
    # Filter only the orders from NSE
    return [order for order in orders if order.get('Exchange') == 'NSE']

    if broker == "aliceblue":
        print(f"Fetching equity orders for {user} from Alice Blue")
        orders = fetch_aliceblue_orders(credentials["username"], credentials["api_key"])

        processed_orders = {}
        for order in orders:
            tradingsymbol = order.get('Symbol', 'N/A')
            transaction_type = order.get('transaction_type', 'N/A')

            # Process Buy and Sell orders
            if transaction_type == "BUY":
                processed_orders[tradingsymbol] = {
                    'Entry Time': order.get('order_timestamp', 'N/A'),
                    'Entry Price': float(order.get('Buyavgprc', 0)),
                    'Qty': float(order.get('Bqty', 0))
                }
            elif transaction_type == "SELL" and tradingsymbol in processed_orders:
                processed_orders[tradingsymbol].update({
                    'Exit Time': order.get('order_timestamp', 'N/A'),
                    'Exit Price': float(order.get('average_price', 0))
                })

        workbook_path = os.path.join(script_dir, "excel", f"{user}.xlsx")
        workbook = load_workbook(workbook_path)
        sheet = workbook["Holdings"]
        last_row = find_first_empty_row(sheet)

        for tradingsymbol, order_data in processed_orders.items():
            if is_order_present(sheet, tradingsymbol, order_data['Entry Time']):
                print(f"Order for {tradingsymbol} at {order_data['Entry Time']} already present. Skipping...")
                continue

            # Calculating data for each row
            entry_price = order_data['Entry Price']
            qty = order_data['Qty']
            exit_price = order_data.get('Exit Price', 0)
            trade_points = exit_price - entry_price
            pnl = trade_points * qty
            margin_used = entry_price * qty  # Both are float values

            row_data = [last_row-1, tradingsymbol, order_data['Entry Time'], order_data.get('Exit Time', 'N/A'), entry_price, exit_price, trade_points, qty, pnl, margin_used]

            # Writing data to Excel sheet
            for col_num, data in enumerate(row_data, 1):
                sheet.cell(row=last_row, column=col_num, value=data)

            last_row += 1

        workbook.save(workbook_path)
