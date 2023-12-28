import os
import sys
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from babel.numbers import format_currency
from telethon.sync import TelegramClient
from dotenv import load_dotenv

# Setting up directory paths
DIR = os.getcwd()
sys.path.append(DIR)

# Import custom modules
import MarketUtils.general_calc as general_calc
import MarketUtils.Main.dtdautomation as dtd

# Get the script directory and file paths
broker_filepath = os.path.join(DIR, "MarketUtils", "broker.json")
ENV_PATH = os.path.join(DIR, '.env')

# Loading environment variables from .env file
load_dotenv(ENV_PATH)
excel_dir = r"C:\Users\vanis\OneDrive\DONOTTOUCH\excel"
# excel_dir = os.getenv('onedrive_excel_folder')
# excel_dir = os.path.join(DIR, "UserProfile","excel")
api_id = os.getenv('telethon_api_id')
api_hash = os.getenv('telethon_api_hash')

# Function to load an existing Excel file and return its data as a dictionary of pandas DataFrames
def load_existing_excel(excel_path):
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    try:
        book = load_workbook(excel_path)
        return {sheet_name: pd.read_excel(excel_path, sheet_name=sheet_name) for sheet_name in book.sheetnames}
    except Exception as e:
        print(f"An error occurred while loading the Excel file: {excel_path}")
        print("Error:", e)
        return {}

def update_json_data(data, user, net_pnl, current_capital,expected_capital, broker_filepath):
    for username in data:
        if user["account_name"] == username["account_name"]:
            user_details = username
            user_details["current_capital"] = round(current_capital, 2)
            user_details["yesterday_PnL"] = round(net_pnl,2)
            user_details["expected_morning_balance"] = round(expected_capital, 2)
    general_calc.write_json_file(broker_filepath,data )

# Function to construct a PNL message for the user
def build_message(user, strategy_results, gross_pnl, total_tax, current_capital, expected_capital):
    message_parts = [
        f"Hello {user}, We hope you're enjoying a wonderful day.\n Here are your PNLs for today:\n"
    ]

       # Iterating over each trade_id and its results
    for trade_id, values in strategy_results.items():
        formatted_pnl = format_currency(values['pnl'], 'INR', locale='en_IN')
        # Using trade_id in the message instead of the strategy name
        message_parts.append(f"{trade_id}: {formatted_pnl}")

    message_parts.extend([
        f"\nGross PnL: {format_currency(gross_pnl, 'INR', locale='en_IN')}",
        f"Expected Tax: {format_currency(total_tax, 'INR', locale='en_IN')}",
        f"Current Capital: {format_currency(current_capital, 'INR', locale='en_IN')}",
        f"Expected Morning Balance : {format_currency(expected_capital, 'INR', locale='en_IN')}",
        "\nBest Regards,\nSerendipity Trading Firm"
    ])

    return "\n".join(message_parts).replace('\u20b9', '₹')


# Function to send a message via Telegram
def send_telegram_message(phone_number, message):
    # Define the session file path
    session_filepath = os.path.join(DIR, "MarketUtils", "Telegram", "+918618221715.session")
    
    # Create a Telegram client and send the message
    with TelegramClient(session_filepath, api_id, api_hash) as client:
        client.send_message(phone_number, message, parse_mode='md')


# The main function which processes user data and sends PNL messages
def main():
    # Load broker data from JSON file
    broker_data = general_calc.read_json_file(broker_filepath)
    
    # Filter active users
    active_users = [user for user in broker_data if "Active" in user['account_type']]
    # phone_number = user["mobile_number"]

    # Get the current date
    today = datetime.now().strftime('%Y-%m-%d')

    for user in active_users:
        excel_path = os.path.join(excel_dir, f"{user['account_name']}.xlsx")
        all_dfs = load_existing_excel(excel_path)
        
        # Update the DTD sheet in the loaded Excel file
        # dtd.main() 

        # Create a dictionary to store the aggregated results for each strategy
        strategy_results = {}
        
        # Initialize total_tax, total_pnl, and pnl_sum to store the cumulative values from all sheets
        total_tax = 0
        total_pnl = 0

        for sheet_name, df in all_dfs.items():  
            if 'exit_time' in df.columns and 'trade_id' in df.columns:
                # df['entry_time'] = pd.to_datetime(df['entry_time']).dt.strftime('%Y-%m-%d')
                df['exit_time'] = pd.to_datetime(df['exit_time']).dt.strftime('%Y-%m-%d')

                df_today = df[df['exit_time'] == today]

                if not df_today.empty:
                    # Calculate the sum of taxes for the trades of the day
                    tax_sum = round(df_today['tax'].sum(), 2)
                    total_tax += tax_sum  # Update the total tax

                    # Iterate through each trade in today's dataframe
                    for index, trade in df_today.iterrows():
                        trade_id = trade['trade_id']
                        pnl = trade['pnl']

                        # Initialize or update the pnl for each trade_id
                        if trade_id not in strategy_results:
                            strategy_results[trade_id] = {'pnl': 0, 'tax': tax_sum}
                        strategy_results[trade_id]['pnl'] += pnl

                        # Update the total pnl
                        total_pnl += pnl

        # Calculate net pnl and expected capital
        gross_pnl = total_pnl   
        expected_tax = total_tax
        net_pnl = gross_pnl - expected_tax
        current_capital = next(account["current_capital"] for account in broker_data if account["account_name"] == user["account_name"])
        expected_capital = current_capital + net_pnl if net_pnl > 0 else current_capital - abs(net_pnl)

        message = build_message(user['account_name'], strategy_results, gross_pnl, expected_tax, current_capital, expected_capital)
        print(message)
        # update_json_data(broker_data, user, net_pnl, current_capital, expected_capital, broker_filepath)

        # # Sending the message via Telegram is currently commented out, remove the comments to enable.
        # send_telegram_message(user['mobile_number'], message)


# Execute the main function when the script is run
if __name__ == "__main__":
    main()