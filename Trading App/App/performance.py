import firebase_admin
from firebase_admin import credentials, storage
from io import BytesIO
import openpyxl
import streamlit as st
from streamlit_calendar import calendar
from streamlit_option_menu import option_menu

# Initialize Firebase app
if not firebase_admin._apps:
    cred = credentials.Certificate("credentials.json")
    firebase_admin.initialize_app(cred, {
        'databaseURL': 'https://trading-app-caf8e-default-rtdb.firebaseio.com',
        'storageBucket': 'trading-app-caf8e.appspot.com'
    })

# Initialize variables
data = []  # This will hold the Excel data

# CSS style definitions
selected = option_menu(None, ["Calendar", "Statistics", "Graph"],
                       icons=['calendar', 'file-bar-graph', 'graph-up'],
                       menu_icon="cast", default_index=0, orientation="horizontal",
                       styles={
                           "container": {"padding": "0!important", "background-color": "#fafafa"},
                           "icon": {"color": "orange", "font-size": "25px"},
                           "nav-link": {"font-size": "25px", "text-align": "left", "margin": "0px", "--hover-color": "#eee"},
                           "nav-link-selected": {"background-color": "purple"},
}
)

# Calendar functionality
if selected == "Calendar":
    calendar_options = {
        "editable": "true",
        "navLinks": "true",
        "initialView": "multiMonthYear"
    }

    # Assuming calendar() returns only the selected date
    selected_date = calendar(
        options=calendar_options,
        key="multiMonthYear",
    )

    # If a date is selected, filter the Excel data
    if selected_date:
        filtered_data = [
            record for record in data if record[2] == selected_date]

        # Display the filtered data
        for record in filtered_data:
            print(record)

# Reference the Firebase Storage bucket
bucket = storage.bucket()

# List all files in the bucket to verify if 'brijesh.xlsx' exists
blobs = bucket.list_blobs()
file_exists = False
for blob in blobs:
    if blob.name == 'brijesh.xlsx':
        file_exists = True
        break

# Proceed only if the file exists
if file_exists:
    # Reference the specific blob (file) in the bucket
    blob = bucket.blob('brijesh.xlsx')

    # Download the blob into an in-memory bytes object
    byte_stream = BytesIO()
    blob.download_to_file(byte_stream)
    byte_stream.seek(0)

    # Load the Excel workbook from the bytes object
    wb = openpyxl.load_workbook(byte_stream)

    # Loop through each sheet in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"Extracting data from sheet: {sheet_name}")

        # Get column names and their indices from the first row
        column_indices = {}
        for idx, cell in enumerate(sheet[1]):
            column_indices[cell.value] = idx

        # Loop through each row in the sheet to read specific columns
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            strategy = row[column_indices['Strategy']].value
            index = row[column_indices['Index']].value
            date = row[column_indices['Date']].value.strftime(
                '%Y-%m-%d')  # Format date from Excel
            pnl = row[column_indices['PnL']].value

            # Append the row data to the list
            data.append([strategy, index, date, pnl])
else:
    st.error("The file 'brijesh.xlsx' does not exist in the bucket.")

# Display the calendar for date selection
selected_date = st.date_input("Select a Date")

# If a date is selected, filter and display the data
if selected_date:
    filtered_data = [record for record in data if record[2]
                     == selected_date.strftime('%Y-%m-%d')]
    if filtered_data:
        st.write("Selected Date Data:")
        for record in filtered_data:
            st.write(f"Strategy: {record[0]}, PnL: {record[3]}")
    else:
        st.warning("No data found for the selected date.")
