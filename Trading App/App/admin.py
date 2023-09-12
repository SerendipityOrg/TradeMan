import shutil
from pathlib import Path
import tempfile
import os
import re
import io
from PIL import Image
import datetime
import base64
import pandas as pd
from firebase_admin import db
from firebase_admin import credentials, storage
import firebase_admin
import hashlib
import openpyxl
from io import BytesIO
import streamlit as st
from streamlit_calendar import calendar
from streamlit_option_menu import option_menu


# Initialize session_state if it doesn't exist
if 'client_data' not in st.session_state:
    st.session_state.client_data = {}


table_style = """
<style>
table.dataframe {
    border-collapse: collapse;
    width: 100%;
}

table.dataframe th,
table.dataframe td {
    border: 1px solid black;
    padding: 8px;
    text-align: left; /* Align text to the left */
}

table.dataframe th {
    background-color: #f2f2f2;
}

table.dataframe tr:nth-child(even) {
    background-color: #f2f2f2;
}

table.dataframe tr:hover {
    background-color: #ddd;
}
</style>
"""

# Initialize Firebase app
if not firebase_admin._apps:
    # Initialize Firebase app
    cred = credentials.Certificate("credentials.json")
    firebase_admin.initialize_app(cred, {
        'databaseURL': 'https://trading-app-caf8e-default-rtdb.firebaseio.com',
        'storageBucket': 'trading-app-caf8e.appspot.com'
    })

    # Initialize variables
data = []  # This will hold the Excel data



def login_admin(username, password):
    hashed_password = hashlib.sha256(password.encode()).hexdigest()

    # Get a reference to the Firebase database
    ref = db.reference('/admin')

    # Check if the username and password match an existing admin
    existing_admins = ref.order_by_child('username').equal_to(username).get()

    if not existing_admins:
        return False
    else:
        for admin_id, admin_data in existing_admins.items():
            if admin_data.get('password') == hashed_password:
                return True
        return False

def get_weeks_for_month(month_number, year):
    first_day_of_month = datetime.date(year, month_number, 1)
    last_day_of_month = datetime.date(year, month_number + 1, 1) - datetime.timedelta(days=1)
    
    # Find the first Saturday of the month
    while first_day_of_month.weekday() != 5:
        first_day_of_month += datetime.timedelta(days=1)
    
    weeks = []
    while first_day_of_month <= last_day_of_month:
        end_day = first_day_of_month + datetime.timedelta(days=6)
        if end_day > last_day_of_month:
            end_day = last_day_of_month
        weeks.append((first_day_of_month.day, end_day.day))
        first_day_of_month += datetime.timedelta(days=7)
    
    return weeks

def update_client_data(client_name, updated_data):
    # Get a reference to the selected client's database
    selected_client_ref = db.reference(f"/clients/{client_name}")
    # Update the client data in the Firebase database
    selected_client_ref.update(updated_data)


def update_brokers_data(client_name, brokers_list_1, brokers_list_2):
    # Get a reference to the selected client's brokers list in the Firebase database
    selected_client_ref = db.reference(f"/clients/{client_name}")
    # Update the brokers list in the Firebase database
    selected_client_ref.update(
        {"Brokers list 1": brokers_list_1, "Brokers list 2": brokers_list_2})


def update_strategies_data(client_name, strategies_list):
    # Get a reference to the selected client's strategies list in the Firebase database
    selected_client_ref = db.reference(f"/clients/{client_name}/Strategy list")
    # Update the strategies list in the Firebase database
    selected_client_ref.set(strategies_list)


def update_profile_picture(selected_client_name, new_profile_picture):
    # Get a reference to the selected client's database
    selected_client_ref = db.reference(f"/clients/{selected_client_name}")

    # Save the uploaded file to a temporary location
    temp_dir = Path(tempfile.gettempdir())
    temp_file_path = temp_dir / new_profile_picture.name

    # Write the uploaded file's content to the temporary location
    with open(temp_file_path, 'wb') as out_file:
        shutil.copyfileobj(new_profile_picture, out_file)

    # Read the temporary file as bytes
    with open(temp_file_path, 'rb') as file:
        profile_picture_bytes = file.read()

    # Convert the image to base64
    profile_picture = base64.b64encode(profile_picture_bytes).decode('utf-8')

    # Update the profile picture in the Firebase database
    selected_client_ref.update({"Profile Picture": profile_picture})

    # Remove the temporary file
    temp_file_path.unlink()


def select_client():
    client_ref = db.reference('/clients')  # Reference to the client database
    client_data = client_ref.get()  # Retrieve the client data

    # Create a list of client names for the select box
    client_names = ['Select'] + list(client_data.keys())

    # Modify client names: replace underscores with spaces and capitalize the first letter of each word
    formatted_client_names = [client_name.replace(
        '_', ' ').title() for client_name in client_names]

    # Select box to choose a client
    selected_client_name = st.sidebar.selectbox(
        'Select a client', formatted_client_names)

    if selected_client_name == 'Select':
        return  # Return early if no client is selected

    # Convert selected_client_name back to original format (with underscores)
    original_selected_client_name = selected_client_name.replace(
        ' ', '_').lower()

    # Check if the selected client name is a valid key in client_data
    if original_selected_client_name not in client_data:
        st.sidebar.warning("Selected client data not found.")
        return

    # Find the selected client's data
    selected_client = client_data[original_selected_client_name]

    # After client is selected, show the next selectbox
    next_selection = st.sidebar.selectbox(
        'Client Details', ['Profile', 'Performance Dashboard'])

    # If 'Profile' is selected, proceed with displaying profile information
    if next_selection == 'Profile':
        display_profile(selected_client, selected_client_name)

    # If 'Performance Dashboard' is selected, display the dashboard info and return
    if next_selection == 'Performance Dashboard':
        display_performance_dashboard(selected_client)


def display_profile(selected_client, selected_client_name):
    profile_picture = selected_client.get("Profile Picture")
    # Display the profile picture if available
    if profile_picture is not None:
        # Decode base64 string to bytes
        profile_picture_bytes = base64.b64decode(profile_picture)

        # Convert profile picture from bytes to PIL Image
        image = Image.open(io.BytesIO(profile_picture_bytes))

       # Convert the image to RGB
        image_rgb = image.convert("RGB")

        # Save the image in JPG format with reduced quality (adjust the quality value as needed)
        image_path = "profile_picture.jpg"
        # Adjust the quality value as needed
        image_rgb.save(image_path, "JPEG", quality=50)

        # Define CSS style to position the profile picture in the right top corner with some margin
        css_style = f"""
            <style>
                .profile-picture-container {{
                    position: absolute;
                    top: -90px;  /* Adjust the top value as needed */
                    right: 10px;
                    border: 2px solid #ccc;
                    border-radius: 50%;
                    overflow: hidden;
                }}
                .profile-picture-container img {{
                    width: 100px;
                    height: 100px;
                }}
            </style>
        """

        # Display the CSS style
        st.markdown(css_style, unsafe_allow_html=True)

        # Display the profile picture in a container with the defined CSS style
        st.markdown(f"""
            <div class="profile-picture-container">
                <img src="data:image/jpeg;base64,{base64.b64encode(profile_picture_bytes).decode('utf-8')}" alt="Profile Picture">
            </div>
        """, unsafe_allow_html=True)

        # Remove the saved image file
        os.remove(image_path)

        # Extract client data from the dictionary
        Name = selected_client.get("Name", "")
        Username = selected_client.get("Username", "")
        Email = selected_client.get("Email", "")
        Password = selected_client.get("Password", "")
        Phone_Number = selected_client.get("Phone Number", "")
        Date_of_Birth = selected_client.get("Date of Birth", "")
        Aadhar_Card_No = selected_client.get("Aadhar Card No", "")
        PAN_Card_No = selected_client.get("PAN Card No", "")
        Bank_Name = selected_client.get("Bank Name", "")
        Bank_Account_No = selected_client.get("Bank Account No", "")
        Brokers_list_1 = selected_client.get("Brokers list 1", [])
        Brokers_list_2 = selected_client.get("Brokers list 2", [])
        Strategy_list = selected_client.get("Strategy list", [])
        Comments = selected_client.get("Comments", "")
        Smart_Contract = selected_client.get("Smart Contract", "")

        # Create a DataFrame to display the client data in tabular form
        data = {
            "Field": ["Name", "Username", "Email", "Password", "Phone Number", "Date of Birth", "Aadhar Card No",
                      "PAN Card No", "Bank Name", "Bank Account No", "Comments", "Smart Contract"],
            "Value": [str(Name), str(Username), str(Email), str(Password), str(Phone_Number), str(Date_of_Birth),
                      str(Aadhar_Card_No), str(PAN_Card_No), str(
                          Bank_Name), str(Bank_Account_No), str(Comments),
                      str(Smart_Contract)]
        }
        df = pd.DataFrame(data)

        # Display the DataFrame as a table with CSS styling and remove index column
        st.markdown(table_style, unsafe_allow_html=True)
        st.write(df.to_html(index=False, escape=False), unsafe_allow_html=True)

        # Display the broker list in a separate table
        st.subheader("Brokers")
        st.write("Broker 1")
        if isinstance(Brokers_list_1, list) and len(Brokers_list_1) > 0:
            broker_1_data = {
                "Field": [],
                "Value": []
            }
            for broker_1 in Brokers_list_1:
                broker_name = broker_1.get("broker_name", "")
        # If broker_name is a list, extract the first element
            if isinstance(broker_name, list) and broker_name:
                broker_name = broker_name[0]

            broker_1_data["Field"].extend(["Broker Name", "User Name", "Password", "2FA",
                                           "TotpAuth", "ApiKey", "ApiSecret", "Active", "Capital", "Risk profile"])
            broker_1_data["Value"].extend([
                str(broker_name),  # Use the processed broker_name value here
                str(broker_1.get("user_name", "")),
                str(broker_1.get("password", "")),
                str(broker_1.get("two_fa", "")),
                str(broker_1.get("totp_auth", "")),
                str(broker_1.get("api_key", "")),
                str(broker_1.get("api_secret", "")),
                str(broker_1.get("active", False)),
                "{:.2f}".format(broker_1.get("capital", 0.00)),
                str(broker_1.get("risk_profile", ""))
            ])

            broker_1_df = pd.DataFrame(broker_1_data)

            # Convert the values in the DataFrame to strings
            broker_1_df = broker_1_df.astype(str)

            # Display the DataFrame as a table with CSS styling and remove index column
            st.markdown(table_style, unsafe_allow_html=True)
            st.write(broker_1_df.to_html(index=False, escape=False),
                     unsafe_allow_html=True)

            # Add some space between the table and "Broker 2"
            st.markdown("<br>", unsafe_allow_html=True)  # Add this line

        st.write("Broker 2")
        if isinstance(Brokers_list_2, list) and len(Brokers_list_2) > 0:
            broker_2_data = {
                "Field": [],
                "Value": []
            }
            for broker_2 in Brokers_list_2:
                broker_name = broker_2.get("broker_name", "")
        # If broker_name is a list, extract the first element
            if isinstance(broker_name, list) and broker_name:
                broker_name = broker_name[0]

            for broker_2 in Brokers_list_2:
                broker_2_data["Field"].extend(["Broker Name", "User Name", "Password", "2FA",
                                               "TotpAuth", "ApiKey", "ApiSecret", "Active", "Capital", "Risk profile"])
                broker_2_data["Value"].extend([
                    str(broker_2.get("user_name", "")),
                    str(broker_2.get("password", "")),
                    str(broker_2.get("two_fa", "")),
                    str(broker_2.get("totp_auth", "")),
                    str(broker_2.get("api_key", "")),
                    str(broker_2.get("api_secret", "")),
                    str(broker_2.get("active", False)),
                    "{:.2f}".format(broker_2.get("capital", 0.00)),
                    str(broker_2.get("risk_profile", ""))
                ])

            broker_2_df = pd.DataFrame(broker_2_data)

            # Convert the values in the DataFrame to strings
            broker_2_df = broker_2_df.astype(str)

            # Display the DataFrame as a table with CSS styling and remove index column
            st.markdown(table_style, unsafe_allow_html=True)
            st.write(broker_2_df.to_html(index=False, escape=False),
                     unsafe_allow_html=True)
        else:
            st.warning("No broker data available.")

        # Display the strategy list in a separate table
        st.subheader("Strategies")
        if isinstance(Strategy_list, list) and len(Strategy_list) > 0:
            strategy_data = {
                "Strategy Name": [],
                "Broker": [],
                "Percentage Allocated": []
            }
        for strategy in Strategy_list:
            strategy_name = strategy.get("strategy_name", "")
            broker = strategy.get("broker", "")

            for selected_strategy in strategy_name:
                for selected_broker in broker:
                    perc_allocated_key = f"strategy_perc_allocated_{selected_strategy}_{selected_broker}_0"
                    percentage_allocated = strategy.get(perc_allocated_key, "")

                    strategy_data["Strategy Name"].append(selected_strategy)
                    strategy_data["Broker"].append(selected_broker)
                    strategy_data["Percentage Allocated"].append(
                        percentage_allocated)

        strategy_df = pd.DataFrame(strategy_data)
        # Display the DataFrame as a table with CSS styling and remove index column
        st.markdown(table_style, unsafe_allow_html=True)
        st.write(strategy_df.to_html(index=False, escape=False),
                 unsafe_allow_html=True)

        # Add some space between the table and the 'Edit' button
        st.markdown("<br>", unsafe_allow_html=True)

        # Add 'Edit' button to switch to 'edit mode'
        if st.button('Edit'):
            st.session_state.edit_mode = True

        # In edit mode, display the editable fields
        if st.session_state.get('edit_mode', False):
            updated_data = {}  # Store the updated client data
            updated_brokers_list_1 = []  # Store the updated brokers list 1
            updated_brokers_list_2 = []  # Store the updated brokers list 2
            updated_strategies_list = []  # Store the updated strategies list

            for i in range(len(df)):
                field = df.at[i, 'Field']
                value = df.at[i, 'Value']
                new_value = st.text_input(
                    field, value=value, key=f"{selected_client_name}-{field}")
                updated_data[field] = new_value

            # Edit profile picture
            st.subheader("Profile Picture")
            # Display the profile picture if available
            if profile_picture is not None or st.session_state.get('edit_mode', False):
                # Decode base64 string to bytes
                if profile_picture is not None:
                    profile_picture_bytes = base64.b64decode(profile_picture)
                else:
                    profile_picture_bytes = b""

                # Convert profile picture from bytes to PIL Image
                image = Image.open(io.BytesIO(profile_picture_bytes))

                # Display the image using st.image
                st.image(image, caption="Profile Picture",
                         use_column_width=True)
            new_profile_picture = st.file_uploader(
                "Upload New Profile Picture", type=["jpg", "jpeg"])
            if new_profile_picture is not None:
                update_profile_picture(
                    selected_client_name, new_profile_picture)
                st.success('Profile picture updated successfully.')
            elif profile_picture is None:
                new_profile_picture = st.file_uploader(
                    "Upload New Profile Picture", type=["jpg", "jpeg"])
                if new_profile_picture is not None:
                    update_profile_picture(
                        selected_client_name, new_profile_picture)

            # Edit brokers list
            st.subheader("Brokers")
            st.write("Broker 1")
            if isinstance(Brokers_list_1, list) and len(Brokers_list_1) > 0:
                for i, broker_1 in enumerate(Brokers_list_1):
                    broker_1["broker_name"] = st.multiselect(
                        "Broker Name", ["Zerodha", "AliceBlue"], default=broker_1.get("broker_name", []),
                        key=f"broker_name_1_{i}")
                    broker_1["user_name"] = st.text_input(
                        "User Name:", key=f"user_name_1_{i}", value=broker_1.get("user_name", ""))
                    broker_1["password"] = st.text_input(
                        "Password:", key=f"password_1_{i}", value=broker_1.get("password", ""))
                    broker_1["two_fa"] = st.text_input(
                        "2FA:", key=f"two_fa_1_{i}", value=broker_1.get("two_fa", ""))
                    broker_1["totp_auth"] = st.text_input(
                        "TotpAuth:", key=f"totp_auth_1_{i}", value=broker_1.get("totp_auth", ""))
                    broker_1["api_key"] = st.text_input(
                        "ApiKey:", key=f"api_key_1_{i}", value=broker_1.get("api_key", ""))
                    broker_1["api_secret"] = st.text_input(
                        "ApiSecret:", key=f"api_secret_1_{i}", value=broker_1.get("api_secret", ""))
                    broker_1["active"] = st.checkbox(
                        "Active:", key=f"active_1_{i}", value=broker_1.get("active", False))
                    broker_1["capital"] = st.number_input(
                        "Capital:", key=f"capital_1_{i}", value=broker_1.get("capital", ""))
                    broker_1["risk_profile"] = st.text_input("Risk profile:", key=f"risk_profile_1_{i}",
                                                             value=broker_1.get("risk_profile", ""))
                    updated_brokers_list_1.append(broker_1)

            st.write("Broker 2")

            if isinstance(Brokers_list_2, list) and len(Brokers_list_2) > 0:
                for i, broker_2 in enumerate(Brokers_list_2):
                    broker_2["broker_name"] = st.multiselect(
                        "Broker Name", ["Zerodha", "AliceBlue"], default=broker_2.get("broker_name", []),
                        key=f"broker_name_2_{i}")
                    broker_2["user_name"] = st.text_input(
                        "User Name:", key=f"user_name_2_{i}", value=broker_2.get("user_name", ""))
                    broker_2["password"] = st.text_input(
                        "Password:", key=f"password_2_{i}", value=broker_2.get("password", ""))
                    broker_2["two_fa"] = st.text_input(
                        "2FA:", key=f"two_fa_2_{i}", value=broker_2.get("two_fa", ""))
                    broker_2["totp_auth"] = st.text_input(
                        "TotpAuth:", key=f"totp_auth_2_{i}", value=broker_2.get("totp_auth", ""))
                    broker_2["api_key"] = st.text_input(
                        "ApiKey:", key=f"api_key_2_{i}", value=broker_2.get("api_key", ""))
                    broker_2["api_secret"] = st.text_input(
                        "ApiSecret:", key=f"api_secret_2_{i}", value=broker_2.get("api_secret", ""))
                    broker_2["active"] = st.checkbox(
                        "Active:", key=f"active_2_{i}", value=broker_2.get("active", False))
                    broker_2["capital"] = st.number_input(
                        "Capital:", key=f"capital_2_{i}", value=broker_2.get("capital", ""))
                    broker_2["risk_profile"] = st.text_input("Risk profile:", key=f"risk_profile_2_{i}",
                                                             value=broker_2.get("risk_profile", ""))
                    updated_brokers_list_2.append(broker_2)

            # Edit strategies list
            st.subheader("Strategies")
            if isinstance(Strategy_list, list) and len(Strategy_list) > 0:
                for i, strategy in enumerate(Strategy_list):
                    strategy["strategy_name"] = st.multiselect(
                        f"Strategy Name {i+1}", ["AmiPy", "MP Wizard", "ZRM", "Overnight Options", "Screenipy Stocks"], default=strategy.get("strategy_name", []), key=f"strategy_name_{i}")
                    strategy["broker"] = st.multiselect(
                        f"Broker {i+1}", ["Zerodha", "AliceBlue"], default=strategy.get("broker", []), key=f"strategy_broker_{i}")
                    selected_strategies = strategy["strategy_name"]
                    selected_brokers = strategy["broker"]

                    for selected_strategy in selected_strategies:
                        for selected_broker_name in selected_brokers:
                            # Modify the key format here
                            perc_allocated_key = f"strategy_perc_allocated_{selected_strategy}_{selected_broker_name}_0"
                            # Retrieve the value using the updated key
                            percentage_allocated = strategy.get(
                                perc_allocated_key, "")
                            options = [f"{i/10:.1f}%" for i in range(0, 101)]
                            default_index = options.index(percentage_allocated)
                            selected_percentage_allocated = st.selectbox(
                                f"Percentage Allocated for {selected_strategy} and {selected_broker_name} (%):", options, index=default_index, key=f"strategy_perc_allocated_{selected_strategy}_{selected_broker_name}_{i}")
                            # Set the selected value
                            strategy[perc_allocated_key] = selected_percentage_allocated
                    updated_strategies_list.append(strategy)

            # Update the changes in the database when Update button is clicked
            if st.button('Update'):
                if updated_data or updated_brokers_list_1 or updated_brokers_list_2 or updated_strategies_list:
                    if updated_data:
                        update_client_data(selected_client_name, updated_data)
                    if updated_brokers_list_1 or updated_brokers_list_2:
                        update_brokers_data(
                            selected_client_name, updated_brokers_list_1, updated_brokers_list_2)
                    if updated_strategies_list:
                        update_strategies_data(
                            selected_client_name, updated_strategies_list)

                    st.success('Client details updated successfully.')
                st.session_state.edit_mode = False  # Switch out of edit mode


# Function to display performance dashboard
def display_performance_dashboard(selected_client):
    # CSS style definitions for the option menu
    selected = option_menu(None, ["Calendar", "Statistics", "Graph"],
                           icons=['calendar', 'file-bar-graph', 'graph-up'],
                           menu_icon="cast", default_index=0, orientation="horizontal",
                           styles={
                               "container": {"padding": "0!important", "background-color": "#fafafa"},
                               "icon": {"color": "orange", "font-size": "25px"},
                               "nav-link": {"font-size": "25px", "text-align": "left", "margin": "0px", "--hover-color": "#eee"},
                               "nav-link-selected": {"background-color": "purple"},
                           })

    # Convert the first letter of client_username to lowercase for file naming
    client_username = selected_client.get("Username", '')
    client_username = client_username[0].lower() + client_username[1:]
    excel_file_name = f"{client_username}.xlsx"  # Construct the Excel file name based on client's username

    # Reference the Firebase Storage bucket
    bucket = storage.bucket('trading-app-caf8e.appspot.com')

    # Check if the client's Excel file exists in the Firebase Storage bucket
    blobs = bucket.list_blobs()
    file_exists = False
    for blob in blobs:
        if blob.name == excel_file_name:
            file_exists = True
            break

    data = []  # List to store extracted data from Excel

    # If the client's Excel file exists, proceed to extract data
    if file_exists:
        # Reference the specific blob (file) in the bucket
        blob = bucket.blob(excel_file_name)

        # Download the blob into an in-memory bytes object
        byte_stream = BytesIO()
        blob.download_to_file(byte_stream)
        byte_stream.seek(0)

        # Load the Excel workbook from the bytes object
        wb = openpyxl.load_workbook(byte_stream, data_only=True)

        # Extract data if the "DTD" sheet exists in the workbook
        if "DTD" in wb.sheetnames:
            sheet = wb["DTD"]
            print(f"Extracting data from sheet: DTD")

            # Get column names and their indices from the first row
            column_indices = {cell.value: idx for idx, cell in enumerate(sheet[1])}

            # Loop through each row in the sheet to read specific columns
            for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
                opening_balance = row[column_indices['Opening Balance']].value
                mp_wizard = row[column_indices['MP Wizard']].value
                date_value = row[column_indices['Date']].value
                date = date_value.strftime('%Y-%m-%d') if date_value else None
                amipy = row[column_indices['AmiPy']].value
                zrm = row [column_indices['ZRM']].value
                overnight_options = row [column_indices['Overnight Options']].value
                gross_pnl = row [column_indices['Gross PnL']].value
                tax = row [column_indices['Tax']].value
                transaction_amount = row[column_indices['Transaction Amount']].value
                deposit_withdrawal = row[column_indices['Deposit/Withdrawal']].value
                # Check if the "Running Balance" column exists in the first row
                if 'Running Balance' in column_indices:
                        running_balance = row[column_indices['Running Balance']].value
                        print(f"Row {row[0].row}: Running Balance value from Excel: {running_balance}")  # Debug print
                else:
                        print("Running Balance column not found!")
                        running_balance = None

                data.append([date, opening_balance, mp_wizard, amipy, zrm, overnight_options, gross_pnl, tax, transaction_amount, deposit_withdrawal, running_balance])
        
        def format_value(value, format_type="normal"):
            print(f"Formatting value: {value}")  # Debug print
            formatted_value = ""
            if value is None:
                formatted_value = "N/A"
            elif isinstance(value, str):
                if value.startswith('='):
                    formatted_value = "Formula"
                else:
                    try:
                        float_value = float(value.replace('₹', '').replace(',', ''))
                        if float_value < 0:
                            formatted_value = f'<span class="negative-value">₹ {float_value:,.2f}</span>'
                        else:
                            formatted_value = f'<span class="positive-value">₹ {float_value:,.2f}</span>'
                    except ValueError:
                        formatted_value = value
            else:
                if value < 0:
                    formatted_value = f'<span class="negative-value">₹ {value:,.2f}</span>'
                else:
                    formatted_value = f'<span class="positive-value">₹ {value:,.2f}</span>'

            # Apply formatting based on format_type
            if format_type == "bold":
                return f"<b>{formatted_value}</b>"
            elif format_type == "italic":
                return f"<i>{formatted_value}</i>"
            else:
                return formatted_value

            
        # Add custom CSS for the table and value colors
        st.markdown("""
        <style>
        .custom-table {
        background-color: #E6E6FA;  # Table background color
        width: 80%;  # Set the table width
        font-size: 20px;  # Increase font size for a bigger table
        margin-left: auto;  # Center the table horizontally
        margin-right: auto;  # Center the table horizontally
            }
        .custom-table td {
        padding: 15px;  # Increase padding for larger cells
        border: 1px solid #ddd;  # Add borders to the cells
        }
        .positive-value {
            color: green;
        }
        .negative-value {
            color: red;
        }
        </style>
        """, unsafe_allow_html=True)
        
        # Calendar functionality
    if selected == "Calendar":
        calendar_options = {
        "editable": "true",
        "navLinks": "true",
        "initialView": "multiMonthYear"
    }
        selected_date = calendar(
        options=calendar_options,
        key="multiMonthYear",
    )

        selected_date = st.date_input("Select a Date")

        if selected_date:
            filtered_data = [record for record in data if record[0] == selected_date.strftime('%Y-%m-%d')]
            print(f"Filtered data for date {selected_date}: {filtered_data}")  # Debug print

            if filtered_data:
                for record in filtered_data:
                    # Create a dictionary to store the labels and values
                    field_names = {
                    "Opening Balance": format_value(record[1], "bold"),
                    "MP Wizard": format_value(record[2], "italic"),
                    "AmiPy": format_value(record[3], "italic"),
                    "ZRM": format_value(record[4], "italic"),
                    "Overnight Options": format_value(record[5], "italic"),
                    "Gross PnL": format_value(record[6], "bold"),
                    "Tax": format_value(record[7]),
                    "Net PnL": format_value(record[8]),
                    "Deposit/Withdrawal": format_value(record[9]),
                    "Running Balance": format_value(record[10], "bold")
                }

                    # Format the filtered data
                    table_data = []
                    for record in filtered_data:
                        # Start from index 1 to skip the "Date"
                        for idx, field in enumerate(field_names, start=1):
                            value = format_value(record[idx])
                            if value != "N/A":
                                table_data.append([field, value])
                    

                    # Display the table without header
                    st.write(pd.DataFrame(table_data, columns=None).to_html(classes='custom-table', header=False, index=False, escape=False), unsafe_allow_html=True)


    if selected == "Statistics":
        # Display date input fields for the user to select the start and end dates
        start_date = st.date_input("Select Start Date", datetime.date(2023, 8, 4))
        end_date = st.date_input("Select End Date")

        # Filter the data based on the selected date range
        filtered_data = [record for record in data if record[0] is not None and start_date.strftime('%Y-%m-%d') <= record[0] <= end_date.strftime('%Y-%m-%d')]

        # Extract relevant data from filtered_data
        opening_balances = [record[1] for record in filtered_data]
        running_balances = [record[10] for record in filtered_data]
        gross_pnls = [record[6] for record in filtered_data]
        transaction_amounts = [record[8] for record in filtered_data]
        deposit_withdrawals = [record[9] for record in filtered_data]

        # Calculate statistics
        initial_capital = opening_balances[0] if opening_balances else 0
        ending_capital = running_balances[-1] if running_balances else 0
        if initial_capital is not None and ending_capital is not None:
            net_profit = ending_capital - initial_capital
        else:
            net_profit = 0

        net_profit_percent = (net_profit / initial_capital) * 100 if initial_capital != 0 else 0
        total_profit = sum(gross_pnls)
        avg_profit = total_profit / len(gross_pnls) if gross_pnls else 0
        avg_profit_percent = (avg_profit / initial_capital) * 100 if initial_capital != 0 else 0
        total_commission = sum(transaction_amounts)/2
        total_deposits = sum([amount for amount in deposit_withdrawals if amount is not None and amount > 0])
        total_withdrawal = sum([amount for amount in deposit_withdrawals if amount is not None and amount < 0])

        def format_stat_value(value):
            if value is None:
                return "N/A"
            elif isinstance(value, str):
                if "₹" in value and "-" in value:
                    return f'<span class="negative-value">{value}</span>'
                elif "₹" in value:
                    return f'<span class="positive-value">{value}</span>'
                elif "%" in value and "-" in value:
                    return f'<span class="negative-value">{value}</span>'
                elif "%" in value:
                    return f'<span class="positive-value">{value}</span>'
                else:
                    return value
            elif value < 0:
                return f'<span class="negative-value">₹ {value:,.2f}</span>'
            else:
                return f'<span class="positive-value">₹ {value:,.2f}</span>'

        # Create a DataFrame for the statistics
        stats_data = {
            "Metric": ["Initial Capital", "Ending Capital", "Net Profit", "Net Profit %", "Total Profit", "Avg. Profit", "Avg. Profit %", "Total Commission", "Total Deposits", "Total Withdrawal"],
            "Value": [format_stat_value(initial_capital), format_stat_value(ending_capital), format_stat_value(net_profit), format_stat_value(f"{net_profit_percent:.2f}%"), format_stat_value(total_profit), format_stat_value(avg_profit), format_stat_value(f"{avg_profit_percent:.2f}%"), format_stat_value(total_commission), format_stat_value(total_deposits), format_stat_value(total_withdrawal)]
        }

        stats_df = pd.DataFrame(stats_data)

        # Display the table without index and without column headers, and with custom styles
        st.write(stats_df.to_html(index=False, header=False, classes='custom-table', escape=False), unsafe_allow_html=True)

def login():

    username = st.text_input('Admin Username')
    password = st.text_input('Password', type='password')

    if st.button('Login'):
        if login_admin(username, password):
            st.session_state.login_successful = True
            st.experimental_rerun()
        else:
            st.error('Invalid username or password.')


def logout():
    st.session_state.login_successful = False


def main():
    if st.session_state.get('login_successful', False):
        # If the admin is logged in, show the client selection page
        select_client()
        if st.sidebar.button('Logout'):
            logout()
    else:
        # If the admin is not logged in, show the login page
        login()


if __name__ == '__main__':
    main()