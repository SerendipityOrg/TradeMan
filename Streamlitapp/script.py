import io
import os
import base64
import datetime
from PIL import Image
import streamlit as st
import pandas as pd
from firebase_admin import db
from firebase_admin import credentials, storage
import openpyxl
import json
from io import BytesIO
import pygwalker as pyg
import streamlit.components.v1 as components
from formats import format_value, format_stat_value, indian_format, custom_format
from streamlit_option_menu import option_menu


storage_bucket = os.getenv('STORAGE_BUCKET')

def display_profile_picture(client_data, style=None):
    # Ensure that client_data is a dictionary
    if isinstance(client_data, str):
        # Try to load the string as a JSON dictionary
        try:
            client_data = json.loads(client_data)
        except json.JSONDecodeError:
            return

    # Fetch the profile picture from the client_data dictionary
    profile_picture = client_data.get("Profile Picture")

    # Display the profile picture if available
    if profile_picture is not None:
        # Decode base64 string to bytes
        profile_picture_bytes = base64.b64decode(profile_picture)

        # Convert profile picture from bytes to PIL Image
        image = Image.open(io.BytesIO(profile_picture_bytes))

        # Convert the image to RGB
        image_rgb = image.convert("RGB")

        # Save the image in JPG format with reduced quality (adjust the quality value as needed)
        image_path = "profile_picture.jpg"
        image_rgb.save(image_path, "JPEG", quality=50)

        # Define default CSS style
        css_style_default = {
            "container": {
                "position": "absolute",
                "top": "-80px",
                "right": "-80px",
                "border": "2px solid #ccc",
                "border-radius": "50%",
                "overflow": "hidden"
            },
            "img": {
                "width": "100px",
                "height": "100px"
            }
        }

        # If additional styles are provided, update the default style
        if style:
            for key, value in style.items():
                if key in css_style_default:
                    css_style_default[key].update(value)

        # Convert the style dictionary to a CSS string
        css_style_string = """
            <style>
                .profile-picture-container {{
                    {container_styles}
                }}
                .profile-picture-container img {{
                    {img_styles}
                }}
            </style>
        """.format(
            container_styles='; '.join(
                f'{k}: {v}' for k, v in css_style_default['container'].items()),
            img_styles='; '.join(
                f'{k}: {v}' for k, v in css_style_default['img'].items())
        )

        # Display the CSS style
        st.markdown(css_style_string, unsafe_allow_html=True)

        # Display the profile picture in a container with the defined CSS style
        st.markdown(f"""
            <div class="profile-picture-container">
                <img src="data:image/jpeg;base64,{base64.b64encode(profile_picture_bytes).decode('utf-8')}" alt="Profile Picture">
            </div>
        """, unsafe_allow_html=True)

        # Remove the saved image file
        os.remove(image_path)

def show_profile(client_data):
    # Display profile picture
    display_profile_picture(client_data)

    pd.options.display.float_format = '{:,.2f}'.format
    # Set the title for the Streamlit app
    st.markdown("<h3 style='color: darkblue'>Profile</h3>",
                unsafe_allow_html=True)

    # Extract client data from the dictionary
    Name = client_data.get("Name", "")
    Username = client_data.get("Username", "")
    Email = client_data.get("Email", "")
    Password = client_data.get("Password", "")
    Phone_Number = client_data.get("Phone Number", "")
    Date_of_Birth = client_data.get("Date of Birth", "")
    Aadhar_Card_No = client_data.get("Aadhar Card No", "")
    PAN_Card_No = client_data.get("PAN Card No", "")
    Bank_Name = client_data.get("Bank Name", "")
    Bank_Account_No = client_data.get("Bank Account No", "")
    Brokers_list_1 = client_data.get("Brokers list 1", [])
    Brokers_list_2 = client_data.get("Brokers list 2", [])
    Strategy_list = client_data.get("Strategy list", [])
    weekly_saturday_capital = client_data.get("Weekly Saturday Capital", "")
    Comments = client_data.get("Comments", "")
    Smart_Contract = client_data.get("Smart Contract", "")
    

    # Create a DataFrame to display the client data in tabular form
    data = {
        "Field": ["Name", "Username", "Email", "Password", "Phone Number", "Date of Birth", "Aadhar Card No",
                  "PAN Card No", "Bank Name", "Bank Account No", "Comments", "Smart Contract"],
        "Value": [str(Name), str(Username), str(Email), str(Password), str(Phone_Number), str(Date_of_Birth),
                  str(Aadhar_Card_No), str(PAN_Card_No), str(
                      Bank_Name), str(Bank_Account_No), str(Comments),
                  str(Smart_Contract)]
    }
    df = pd.DataFrame(data)
    # Display the DataFrame as a table with CSS styling and remove index column
    st.markdown(table_style, unsafe_allow_html=True)
    st.write(df.to_html(index=False, escape=False), unsafe_allow_html=True)

    # Display the broker list in vertical tabular form
    st.subheader("Brokers")
    st.write("Broker 1")
    if isinstance(Brokers_list_1, list) and len(Brokers_list_1) > 0:
        broker_1_data = {
            "Field": [],
            "Value": []
        }
        for broker_1 in Brokers_list_1:
            broker_1_data["Field"].extend(["Broker Name", "User Name", "Password", "2FA",
                                           "TotpAuth", "ApiCode", "ApiKey", "ApiSecret", "Active", "Capital", "Risk profile"])
            broker_1_data["Value"].extend([
                str(broker_1.get("broker_name", [""])[0]),
                str(broker_1.get("user_name", "")),
                str(broker_1.get("password", "")),
                str(broker_1.get("two_fa", "")),
                str(broker_1.get("totp_auth", "")),
                str(broker_1.get("api_code", "")),
                str(broker_1.get("api_key", "")),
                str(broker_1.get("api_secret", "")),
                str(broker_1.get("active", "")),
                "{:.2f}".format(broker_1.get("capital", 0.00)),
                str(broker_1.get("risk_profile", ""))
            ])

        broker_1_df = pd.DataFrame(broker_1_data)
        # Display the DataFrame as a table with CSS styling and remove index column
        st.markdown(table_style, unsafe_allow_html=True)
        st.write(broker_1_df.to_html(index=False, escape=False),
                 unsafe_allow_html=True)

        # Add some space between the table and "Broker 2"
        st.markdown("<br>", unsafe_allow_html=True)  # Add this line
    else:
        st.warning("No broker data available.")

    st.write("Broker 2")
    if isinstance(Brokers_list_2, list) and len(Brokers_list_2) > 0:
        broker_2_data = {
            "Field": [],
            "Value": []
        }
        for broker_2 in Brokers_list_2:
            broker_2_data["Field"].extend(["Broker Name", "User Name", "Password", "2FA",
                                           "TotpAuth", "Apicode", "ApiKey", "ApiSecret", "Active", "Capital", "Risk profile"])
            broker_2_data["Value"].extend([
                str(broker_2.get("broker_name", [""])[0]),
                str(broker_2.get("user_name", "")),
                str(broker_2.get("password", "")),
                str(broker_2.get("two_fa", "")),
                str(broker_2.get("totp_auth", "")),
                str(broker_2.get("api_code", "")),
                str(broker_2.get("api_key", "")),
                str(broker_2.get("api_secret", "")),
                str(broker_2.get("active", "")),
                "{:.2f}".format(broker_2.get("capital", 0.00)),
                str(broker_2.get("risk_profile", ""))
            ])

        broker_2_df = pd.DataFrame(broker_2_data)
        # Display the DataFrame as a table with CSS styling and remove index column
        st.markdown(table_style, unsafe_allow_html=True)
        st.write(broker_2_df.to_html(index=False, escape=False),
                 unsafe_allow_html=True)
    else:
        st.warning("No broker data available.")

     # Display the strategy list in vertical tabular form
    st.subheader("Strategies")
    if isinstance(Strategy_list, list) and len(Strategy_list) > 0:
        strategy_data = {
            "Strategy Name": [],
            "Broker": [],
            "Percentage Allocated": []
        }
        for strategy in Strategy_list:
            strategy_name = strategy.get("strategy_name", "")
            broker = strategy.get("broker", "")

            for selected_strategy in strategy_name:
                for selected_broker in broker:
                    perc_allocated_key = f"strategy_perc_allocated_{selected_strategy}_{selected_broker}_0"
                    percentage_allocated = strategy.get(perc_allocated_key, "")

                    strategy_data["Strategy Name"].append(selected_strategy)
                    strategy_data["Broker"].append(selected_broker)
                    strategy_data["Percentage Allocated"].append(
                        percentage_allocated)

        strategy_df = pd.DataFrame(strategy_data)
        # Display the DataFrame as a table with CSS styling and remove index column
        st.markdown(table_style, unsafe_allow_html=True)
        st.write(strategy_df.to_html(index=False, escape=False),
                 unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)  
    
    # Display Week Saturday Capital
    st.subheader("Week Saturday Capital")
    # Initialize week_saturday_capital_value with a default value
    weekly_saturday_capital_value = 0.0

    if weekly_saturday_capital:
        # Ensure week_saturday_capital is a float, default to 0.0 if not
        weekly_saturday_capital_value = float(weekly_saturday_capital) if isinstance(weekly_saturday_capital, (float, int)) else 0.0

    # Calculate the upcoming Saturday date
    today = datetime.date.today()
    next_saturday = today + datetime.timedelta((5 - today.weekday()) % 7)  # 5 represents Saturday
    formatted_saturday = next_saturday.strftime('%d-%m-%Y')  # Format the date as dd-mm-yyyy

    if weekly_saturday_capital:
        # Ensure weekly_saturday_capital is a float, default to 0.0 if not
        weekly_saturday_capital_value = float(weekly_saturday_capital) if isinstance(weekly_saturday_capital, (float, int)) else 0.0

    # Use the custom format function to format the capital value
    formatted_capital_value = custom_format(weekly_saturday_capital_value)

    st.write(f"Weekly Saturday Capital for {formatted_saturday}: {formatted_capital_value}")

table_style = """
<style>
table.dataframe {
    border-collapse: collapse;
    width: 100%;
}

table.dataframe th,
table.dataframe td {
    border: 1px solid black;
    padding: 8px;
    text-align: left; /* Align text to the left */
}

/* Header background color */
table.dataframe th {
    background-color: AliceBlue;
}

/* Alternating row background colors */
table.dataframe tr:nth-child(even) {
    background-color: AliceBlue;  /* Even rows will be AliceBlue */
}
table.dataframe tr:nth-child(odd) {
    background-color: AliceBlue;  /* Odd rows will be white */
}

/* Hover style for rows */
table.dataframe tr:hover {
    background-color: HoneyDew;
}
</style>
"""

# Function to display performance dashboard
def display_performance_dashboard(selected_client, client_data, excel_file_name):
    # Initialize Streamlit page config and title if this is your main entry function
    # st.set_page_config(page_title="Performance Dashboard", layout="wide")
    st.subheader("Performance Dashboard")

    # Reference the Firebase Storage bucket
    bucket = storage.bucket(storage_bucket)

    # Check if the client's Excel file exists in the Firebase Storage bucket
    blobs = bucket.list_blobs()
    file_exists = False
    for blob in blobs:
        if blob.name == excel_file_name:
            file_exists = True
            break

    data = []  # List to store extracted data from Excel

    # If the client's Excel file exists, proceed to extract data
    if file_exists:
        # Reference the specific blob (file) in the bucket
        blob = bucket.blob(excel_file_name)

        # Download the blob into an in-memory bytes object
        byte_stream = BytesIO()
        blob.download_to_file(byte_stream)
        byte_stream.seek(0)

        # Load the Excel workbook from the bytes object
        wb = openpyxl.load_workbook(byte_stream, data_only=True)

        # Extract data if the "DTD" sheet exists in the workbook
        if "DTD" in wb.sheetnames:
            sheet = wb["DTD"]
            print(f"Extracting data from sheet: DTD")

            # Get column names and their indices from the first row
            column_indices = {cell.value: idx for idx, cell in enumerate(sheet[1])}

            # Loop through each row in the sheet to read specific columns
            # Assuming headers are in the first row
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                date = row[column_indices['Date']].value
                day = row[column_indices['Day']].value
                trade_id = row[column_indices['Trade ID']].value
                details = row[column_indices['Details']].value
                amount = row[column_indices['Amount']].value

                # Add the row data to the data list
                data.append([date, day, trade_id, details, amount])

        # Convert the data list to a DataFrame
        df = pd.DataFrame(data, columns=['Date', 'Day', 'Trade ID', 'Details', 'Amount'])

        # Generate the HTML using Pygwalker
        pyg_html = pyg.to_html(df)

        # Embed the HTML into the Streamlit app
        components.html(pyg_html, height=1000, scrolling=True)