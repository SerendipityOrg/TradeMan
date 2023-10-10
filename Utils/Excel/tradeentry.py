import os,sys
import json
import pandas as pd
from openpyxl import load_workbook
from babel.numbers import format_currency
import strategy_calc as sc
import firebase_admin
from firebase_admin import credentials, storage
from telethon.sync import TelegramClient
from dotenv import load_dotenv

env_path = os.path.join(os.path.dirname(__file__), '..','..','Brokers', '.env')
print(load_dotenv(env_path))


api_id = os.getenv('telethon_api_id')
api_hash = os.getenv('telethon_api_hash')

script_dir = os.path.dirname(os.path.realpath(__file__))
utils_dir = os.path.join(script_dir, "..")
sys.path.append(utils_dir)
import general_calc as gc
broker_filepath = os.path.join(utils_dir, "broker.json")
# broker_filepath = os.path.join(script_dir,"..", "broker.json")

userprofile_dir = os.path.join(script_dir, "..","..", "UserProfile")
json_dir = os.path.join(userprofile_dir, "json")
excel_dir = os.path.join(userprofile_dir, "excel")


def custom_format(amount):
    formatted = format_currency(amount, 'INR', locale='en_IN')
    return formatted.replace('₹', '₹ ')

def process_strategy_data(user_data, broker, strategy_name, process_func):
    if strategy_name in user_data[broker]["orders"]:
        data = process_func(broker,user_data[broker]["orders"][strategy_name])
        print(data)
        df = pd.DataFrame(data)
        pnl = round(df["PnL"].sum(), 2)
        tax = round(df["Tax"].sum(), 2)
        return df, pnl, tax
    return pd.DataFrame(), 0, 0

def load_existing_excel(excel_path):
    book = load_workbook(excel_path)
    return {sheet_name: pd.read_excel(excel_path, sheet_name=sheet_name) for sheet_name in book.sheetnames}

def save_all_sheets_to_excel(all_dfs, excel_path):
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        for sheet_name, df in all_dfs.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

def build_message(user, mpwizard_pnl, amipy_pnl, overnight_pnl, gross_pnl, tax, current_capital, expected_capital):
    # Construct the message similar to the original script
    message_parts = [f"Hello {user},We hope you're enjoying a wonderful day.\n Here are your PNLs for today:\n"]

    if mpwizard_pnl != 0:
        message_parts.append(f"MPWizard: {custom_format(mpwizard_pnl)}")

    if amipy_pnl != 0:
        message_parts.append(f"AmiPy: {custom_format(amipy_pnl)}")

    if overnight_pnl != 0:
        message_parts.append(f"Overnight Options: {custom_format(overnight_pnl)}")

    message_parts.extend([
        f"\n**Gross PnL: {custom_format(gross_pnl)}**",
        f"**Expected Tax: {custom_format(tax)}**",
        f"**Current Capital: {custom_format(current_capital)}**",
        f"**Expected Morning Balance : {custom_format(expected_capital)}**",
        "\nBest Regards,\nSerendipity Trading Firm"
    ])
    
    return message_parts

def update_json_data(data, broker, user, net_pnl, expected_capital, broker_filepath):
    user_details = data[broker][user]
    user_details["yesterday_PnL"] = net_pnl
    user_details["expected_morning_balance"] = round(expected_capital, 2)
    data[broker][user] = user_details

    with open(broker_filepath, 'w') as json_file:
        json.dump(data, json_file, indent=4)

def update_excel_data(all_dfs, mpwizard_df, amipy_df, overnight_df):
    if not mpwizard_df.empty:
        all_dfs["MPWizard"] = pd.concat([all_dfs.get("MPWizard", pd.DataFrame()), mpwizard_df])
    if not amipy_df.empty:
        all_dfs["AmiPy"] = pd.concat([all_dfs.get("AmiPy", pd.DataFrame()), amipy_df])
    if not overnight_df.empty:
        all_dfs["Overnight_options"] = pd.concat([all_dfs.get("Overnight_options", pd.DataFrame()), overnight_df])


cred = credentials.Certificate("TradeMan/Utils/Excel/credentials.json")
firebase_admin.initialize_app(cred, {
    'databaseURL': 'https://trading-app-caf8e-default-rtdb.firebaseio.com'
})

def save_to_firebase(user, excel_path):
    
    # Correct bucket name
    bucket = storage.bucket(name='trading-app-caf8e.appspot.com')
    blob = bucket.blob(f'{user}.xlsx')
    with open(excel_path, 'rb') as my_file:
        blob.upload_from_file(my_file)
    print(f"Excel file for {user} has been uploaded to Firebase.")

def send_telegram_message(phone_number, message):
    session_filepath = os.path.join(script_dir, "..",'..','..', "+918618221715.session")
    with TelegramClient(session_filepath, api_id, api_hash) as client:
        client.send_message(phone_number, message, parse_mode='md')


def main():
    data = gc.read_json_file(broker_filepath)
    user_list = []
    # Go through each broker
    for broker, broker_data in data.items():
        # Check if 'accounts_to_trade' is in the broker data
        if 'accounts_to_trade' in broker_data:
            # Add each account to the list
            for account in broker_data['accounts_to_trade']:
                user_list.append((broker, account))

    for broker, user in user_list:
        user_data = gc.read_json_file(os.path.join(json_dir, f"{user}.json"))
        phone_number = data[broker][user]["mobile_number"]
        
        mpwizard_df, mpwizard_pnl, mpwizard_tax = process_strategy_data(user_data, broker, "MPWizard", sc.process_mpwizard_trades)
        amipy_df, amipy_pnl, amipy_tax = process_strategy_data(user_data, broker, "AmiPy", sc.process_amipy_trades)  # Implement a function that processes Amipy trades
        overnight_df, overnight_pnl, overnight_tax = process_strategy_data(user_data, broker, "Overnight_Options", sc.process_overnight_options_trades)

        gross_pnl = mpwizard_pnl + amipy_pnl + overnight_pnl
        tax = mpwizard_tax + amipy_tax + overnight_tax
        net_pnl = gross_pnl - tax

        current_capital = data[broker][user]['current_capital']
        expected_capital = current_capital + net_pnl if net_pnl > 0 else current_capital - abs(net_pnl)

        message_parts = build_message(user, mpwizard_pnl, amipy_pnl, overnight_pnl, gross_pnl, tax, current_capital, expected_capital)
        message = "\n".join(message_parts).replace('\u20b9', 'INR')
        print(message)

        update_json_data(data, broker, user, net_pnl, expected_capital, broker_filepath)

        excel_path = os.path.join(excel_dir, f"{user}.xlsx")
        all_dfs = load_existing_excel(excel_path)

        update_excel_data(all_dfs, mpwizard_df, amipy_df, overnight_df)
        save_all_sheets_to_excel(all_dfs, excel_path)

        # Assuming you want to save to Firebase and send messages as in the original script
        save_to_firebase(user, excel_path)  # Existing function
        send_telegram_message(phone_number, message)  # Separate into a function

# Add other necessary helper functions...

if __name__ == "__main__":
    main()
