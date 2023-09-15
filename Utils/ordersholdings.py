# import os
# import json
# from kiteconnect import KiteConnect, KiteTicker
# from openpyxl import load_workbook
# from pya3 import *
# from pprint import pprint

# # File paths
# script_dir = os.path.dirname(os.path.realpath(__file__))
# broker_filepath = os.path.join(script_dir, "broker.json")

# # Load the broker data from the JSON file
# with open(broker_filepath) as file:
#     data = json.load(file)

# # Function to fetch orders from Zerodha
# def fetch_zerodha_orders(api_key, access_token):
#     kite = KiteConnect(api_key=api_key)
#     kite.set_access_token(access_token)
#     try:
#         orders = kite.orders()
#         # Filter only the orders from NSE
#         return [order for order in orders if order['exchange'] == 'NSE']
#     except Exception as e:
#         return [f"Error fetching orders from Zerodha: {str(e)}"]

# # Function to fetch orders from Alice Blue using smartapi-python
# def fetch_aliceblue_orders(username, api_key):
#     alice = Aliceblue(username, api_key)
#     session_id = alice.get_session_id()
#     orders = alice.get_daywise_positions()
#     # Filter only the orders from NSE
#     return [order for order in orders if order['Exchange'] == 'NSE']

# # Function to find the first empty row in a worksheet
# def find_first_empty_row(sheet):
#     for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
#         if all(cell is None for cell in row):
#             return i
#     return sheet.max_row + 1

# # Function to check if an order is already in the worksheet
# def is_order_present(sheet, tradingsymbol, timestamp):
#     for row in sheet.iter_rows(values_only=True):
#         if row:
#             # Assuming tradingsymbol is in the 2nd column and timestamp is in the 3rd column
#             if row[1] == tradingsymbol and row[2] == timestamp:
#                 return True
#     return False

# # Populate user_list with accounts from each broker
# user_list = []
# for broker, broker_data in data.items():
#     if 'accounts_to_trade' in broker_data:
#         for account in broker_data['accounts_to_trade']:
#             user_list.append((broker, account))

# # Iterate over all the broker-user pairs
# for broker, user in user_list:
#     credentials = data[broker][user]

#     if broker == "zerodha":
#         print(f"Fetching equity orders for {user} from Zerodha")
#         orders = fetch_zerodha_orders(
#             credentials["api_key"], credentials["access_token"])
#     elif broker == "aliceblue":
#         print(f"Fetching equity orders for {user} from Alice Blue")
#         orders = fetch_aliceblue_orders(
#             credentials["username"], credentials["api_key"])
#     else:
#         print(f"Unsupported broker: {broker}")
#         continue

#     processed_orders = {}
#     # Process the fetched orders
#     for order in orders:
#         tradingsymbol = order.get('tradingsymbol', 'N/A')
#         transaction_type = order.get('transaction_type', 'N/A')

#         if transaction_type == "BUY":
#             processed_orders[tradingsymbol] = {
#                 'Entry Time': order.get('order_timestamp', 'N/A'),
#                 'Entry Price': float(order.get('average_price', 0)),
#                 'Qty': float(order.get('quantity', 0))
#             }
#         elif transaction_type == "SELL" and tradingsymbol in processed_orders:
#             processed_orders[tradingsymbol]['Exit Time'] = order.get(
#                 'order_timestamp', 'N/A')
#             processed_orders[tradingsymbol]['Exit Price'] = float(
#                 order.get('average_price', 0))

#     # Adjusted workbook_path to point to 'excel' folder
#     workbook_path = os.path.join(script_dir, "excel", f"{user}.xlsx")
#     workbook = load_workbook(workbook_path)
#     sheet = workbook["Holdings"]

#     # Find the first empty row to start writing data
#     last_row = find_first_empty_row(sheet)

#     # Write processed orders to Excel sheet
#     for tradingsymbol, order_data in processed_orders.items():
#         # Check if the order is already present in the Excel sheet
#         if is_order_present(sheet, tradingsymbol, order_data['Entry Time']):
#             print(
#                 f"Order for {tradingsymbol} at {order_data['Entry Time']} already present. Skipping...")
#             continue

#         entry_price = order_data['Entry Price']
#         qty = order_data['Qty']
#         exit_price = order_data.get('Exit Price', 0)

#         # Calculating Trade Points and PnL
#         trade_points = exit_price - entry_price
#         pnl = trade_points * qty
#         margin_used = entry_price * qty  # Ensure both are float values

#         row_data = [last_row-1, tradingsymbol, order_data['Entry Time'], order_data.get(
#             'Exit Time', 'N/A'), entry_price, exit_price, trade_points, qty, pnl, margin_used]

#         for col_num, data in enumerate(row_data, 1):
#             sheet.cell(row=last_row, column=col_num, value=data)

#         last_row += 1

#     workbook.save(workbook_path)

import os
import json
from kiteconnect import KiteConnect
from openpyxl import load_workbook
from pya3 import Aliceblue
from pprint import pprint

# Function to load data from a JSON file
def load_broker_data(filepath):
    with open(filepath) as file:
        return json.load(file)

# Function to fetch equity orders from Zerodha API
def fetch_zerodha_orders(api_key, access_token):
    kite = KiteConnect(api_key=api_key)
    kite.set_access_token(access_token)
    try:
        orders = kite.orders()
        return [order for order in orders if order['exchange'] == 'NSE']
    except Exception as e:
        return [f"Error fetching orders from Zerodha: {str(e)}"]

# Function to fetch equity orders from Alice Blue API
def fetch_aliceblue_orders(username, api_key):
    alice = Aliceblue(username, api_key)
    session_id = alice.get_session_id()
    orders = alice.get_daywise_positions()
    return [order for order in orders if order['Exchange'] == 'NSE']

# Function to find the first empty row in an Excel sheet
def find_first_empty_row(sheet):
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if all(cell is None for cell in row):
            return i
    return sheet.max_row + 1

# Function to check if an order is present in the Excel sheet
def is_order_present(sheet, tradingsymbol, timestamp):
    for row in sheet.iter_rows(values_only=True):
        if row and row[1] == tradingsymbol and row[2] == timestamp:
            return True
    return False

# Function to process orders and extract relevant information
def process_orders(orders):
    processed_orders = {}
    for order in orders:
        tradingsymbol = order.get('tradingsymbol', 'N/A')
        transaction_type = order.get('transaction_type', 'N/A')

        if transaction_type == "BUY":
            processed_orders[tradingsymbol] = {
                'Entry Time': order.get('order_timestamp', 'N/A'),
                'Entry Price': float(order.get('average_price', 0)),
                'Qty': float(order.get('quantity', 0))
            }
        elif transaction_type == "SELL" and tradingsymbol in processed_orders:
            processed_orders[tradingsymbol]['Exit Time'] = order.get('order_timestamp', 'N/A')
            processed_orders[tradingsymbol]['Exit Price'] = float(order.get('average_price', 0))
    return processed_orders

# Function to write order information to an Excel sheet
def write_to_excel(sheet, last_row, tradingsymbol, order_data):
    entry_price = order_data['Entry Price']
    qty = order_data['Qty']
    exit_price = order_data.get('Exit Price', 0)

    trade_points = exit_price - entry_price
    pnl = trade_points * qty
    margin_used = entry_price * qty

    row_data = [last_row - 1, tradingsymbol, order_data['Entry Time'], order_data.get('Exit Time', 'N/A'), 
                entry_price, exit_price, trade_points, qty, pnl, margin_used]

    for col_num, data in enumerate(row_data, 1):
        sheet.cell(row=last_row, column=col_num, value=data)

# Main function to execute the program
def main():
    script_dir = os.path.dirname(os.path.realpath(__file__))  # Get the directory path of the current script
    broker_filepath = os.path.join(script_dir, "broker.json")  # Construct the path to the broker.json file
    data = load_broker_data(broker_filepath)  # Load data from the broker.json file

    user_list = []
    # Extract broker and account information from the loaded data
    for broker, broker_data in data.items():
        if 'accounts_to_trade' in broker_data:
            for account in broker_data['accounts_to_trade']:
                user_list.append((broker, account))

    # Loop through each user and fetch/process orders
    for broker, user in user_list:
        credentials = data[broker][user]

        if broker == "zerodha":
            print(f"Fetching equity orders for {user} from Zerodha")
            orders = fetch_zerodha_orders(credentials["api_key"], credentials["access_token"])
        elif broker == "aliceblue":
            print(f"Fetching equity orders for {user} from Alice Blue")
            orders = fetch_aliceblue_orders(credentials["username"], credentials["api_key"])
        else:
            print(f"Unsupported broker: {broker}")
            continue

        processed_orders = process_orders(orders)

        workbook_path = os.path.join(script_dir, "excel", f"{user}.xlsx")
        workbook = load_workbook(workbook_path)
        sheet = workbook["Holdings"]
        last_row = find_first_empty_row(sheet)

        # Write processed orders to the Excel sheet
        for tradingsymbol, order_data in processed_orders.items():
            if not is_order_present(sheet, tradingsymbol, order_data['Entry Time']):
                write_to_excel(sheet, last_row, tradingsymbol, order_data)
                last_row += 1

        workbook.save(workbook_path)

if __name__ == "__main__":
    main()  # Call the main function when the script is executed
