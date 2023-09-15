import pandas as pd
import json
from openpyxl import load_workbook
import os,io,sys
from telethon.sync import TelegramClient
from calculations.taxcalculation import *
from babel.numbers import format_currency


api_id = '22941664'
api_hash = '2ee02d39b9a6dae9434689d46e0863ca'

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def process_mpwizard_trades(mpwizard_trades):
    if not mpwizard_trades:
        print("No MPWizard trades found.")
        return []
        
    result = []
    # Extracting trade details
    for i in range(len(mpwizard_trades["BUY"])):
        buy_trade = mpwizard_trades["BUY"][i]
        sell_trade = mpwizard_trades["SELL"][i]

        
        if broker == "zerodha":
            charges = zerodha_taxes(buy_trade["qty"], buy_trade["avg_prc"], sell_trade["avg_prc"],1)
        elif broker == "aliceblue":
            charges = aliceblue_taxes(buy_trade["qty"], float(buy_trade["avg_prc"]), float(sell_trade["avg_prc"]),1)
        

        trade_data = {
            "Strategy": "MPWizard",
            "Index": buy_trade["tradingsymbol"][:-12],
            "Strike Prc": buy_trade["tradingsymbol"][-7:-2],
            "Date": pd.to_datetime(buy_trade["timestamp"]).date(),
            "Entry Time": pd.to_datetime(buy_trade["timestamp"]).strftime('%H:%M'),
            "Exit Time": pd.to_datetime(sell_trade["timestamp"]).strftime('%H:%M'),
            "Entry Price": buy_trade["avg_prc"],
            "Exit Price": sell_trade["avg_prc"],
            "Trade points": float(sell_trade["avg_prc"]) - float(buy_trade["avg_prc"]),
            "Qty": buy_trade["qty"],
            "Tax": charges,
            "PnL": (float(sell_trade["avg_prc"]) - float(buy_trade["avg_prc"])) * int(buy_trade["qty"]) 
        }
        result.append(trade_data)
    return result

def custom_format(amount):
    formatted = format_currency(amount,'INR', locale='en_IN')
    return formatted.replace('₹','₹')

def process_short_trades(short_signals, short_cover_signals):
    if len(short_signals) != len(short_cover_signals):
        print("Mismatch in the number of ShortSignal and ShortCoverSignal trades.")
        return []

    result = []
    for i in range(0, len(short_signals), 4):  # Process each group of 4 trades together
        short_signal_group = short_signals[i:i+4]
        short_cover_signal_group = short_cover_signals[i:i+4]

        hedge_entry = sum(float(trade["avg_prc"]) for trade in short_signal_group if trade["trade_type"] == "HedgeOrder") 
        hedge_exit =  sum(float(trade["avg_prc"]) for trade in short_cover_signal_group if trade["trade_type"] == "HedgeOrder")

        entry_price = sum(float(trade["avg_prc"]) for trade in short_signal_group if trade["trade_type"] == "ShortSignal")
        exit_price = sum(float(trade["avg_prc"]) for trade in short_cover_signal_group if trade["trade_type"] == "ShortCoverSignal")
        hedge_price = hedge_exit - hedge_entry
        trade_points = (entry_price - exit_price) + hedge_price
        

        if broker == "zerodha":
            charges = zerodha_taxes(short_signal_group[0]["qty"], entry_price, exit_price,2)
            hedge_charges = zerodha_taxes(short_signal_group[0]["qty"], hedge_entry, hedge_exit,2)
        elif broker == "aliceblue":
            charges = aliceblue_taxes(short_signal_group[0]["qty"], entry_price, exit_price,2)
            hedge_charges = aliceblue_taxes(short_signal_group[0]["qty"], hedge_entry, hedge_exit,2)
        charges = charges + hedge_charges
        
        trade_data = {
            "Strategy": "Nifty Straddle",
            "Index": "NIFTY",
            "Trade Type": "Short",
            "Strike Prc": short_signal_group[0]["strike_price"],
            "Date": pd.to_datetime(short_signal_group[0]["timestamp"]).date(),
            "Entry Time": pd.to_datetime(short_signal_group[0]["timestamp"]).strftime('%H:%M'),
            "Exit Time": pd.to_datetime(short_cover_signal_group[0]["timestamp"]).strftime('%H:%M'),
            "Entry Price": entry_price,
            "Exit Price": exit_price,
            "Hedge Entry": hedge_entry,
            "Hedge Exit": hedge_exit,
            "Trade points": trade_points,
            "Qty": short_signal_group[0]["qty"],
            "PnL": trade_points * int(short_signal_group[0]["qty"]),
            "Tax": charges            
        }
        result.append(trade_data)
    return result

def process_long_trades(long_signals, long_cover_signals):
    if len(long_signals) != len(long_cover_signals):
        print("Mismatch in the number of LongSignal and LongCoverSignal trades.")
        return []

    result = []
    for i in range(0, len(long_signals), 2):  # Process each pair of trades together
        long_signal_pair = long_signals[i:i+2]
        long_cover_signal_pair = long_cover_signals[i:i+2]

        entry_price = sum(float(trade["avg_prc"]) for trade in long_signal_pair)
        exit_price = sum(float(trade["avg_prc"]) for trade in long_cover_signal_pair)
        trade_points = entry_price - exit_price

        if broker == "zerodha":
            charges = zerodha_taxes(long_signal_pair[0]["qty"], entry_price, exit_price,2)
        elif broker == "aliceblue":
            charges = aliceblue_taxes(long_signal_pair[0]["qty"], entry_price, exit_price,2)

        trade_data = {
            "Strategy": "Amipy",
            "Index": "NIFTY",
            "Trade Type": "Long",
            "Strike Prc": long_signal_pair[0]["strike_price"],
            "Date": pd.to_datetime(long_signal_pair[0]["timestamp"]).date(),
            "Entry Time": pd.to_datetime(long_signal_pair[0]["timestamp"]).strftime('%H:%M'),
            "Exit Time": pd.to_datetime(long_cover_signal_pair[0]["timestamp"]).strftime('%H:%M'),
            "Entry Price": entry_price,
            "Exit Price": exit_price,
            "Trade points": trade_points,
            "Qty": long_signal_pair[0]["qty"],
            "Hedge Price": float('nan'),
            "PnL": trade_points * int(long_signal_pair[0]["qty"]),
            "Tax": charges 
        }
        result.append(trade_data)
    return result

def process_overnight_options_trades(overnight_options_trades):
    if not overnight_options_trades:
        print("No Overnight_Options trades found.")
        return []
    
    result = []
    
    # Extracting trade details from Afternoon and Morning
    afternoon_trades = overnight_options_trades.get("Afternoon", [])
    morning_trades = overnight_options_trades.get("Morning", [])
    qty = afternoon_trades[0]["qty"]
    
    if afternoon_trades[0]["direction"] == "BULLISH":
    # Extracting BULLISH trades with strike_price = 0 for both Afternoon and Morning
        future_entry = next((float(trade['avg_prc']) for trade in afternoon_trades if trade['direction'] == 'BULLISH' and trade['strike_price'] == "0"), None)
        future_exit = next((float(trade['avg_prc']) for trade in morning_trades if trade['direction'] == 'BULLISH' and trade['strike_price'] == "0"), None)    
    # Extracting BULLISH trades with strike_price != 0 for both Afternoon and Morning
        option_entry = next((float(trade['avg_prc']) for trade in afternoon_trades if trade['direction'] == 'BULLISH' and trade['strike_price'] != "0"), None)
        option_exit = next((float(trade['avg_prc']) for trade in morning_trades if trade['direction'] == 'BULLISH' and trade['strike_price'] != "0"), None)
    elif afternoon_trades[0]["direction"] == "BEARISH":
    # Extracting BEARISH trades with strike_price = 0 for both Afternoon and Morning
        future_entry = next((float(trade['avg_prc']) for trade in afternoon_trades if trade['direction'] == 'BEARISH' and trade['strike_price'] == "0"), None)
        future_exit = next((float(trade['avg_prc']) for trade in morning_trades if trade['direction'] == 'BEARISH' and trade['strike_price'] == "0"), None)
    # Extracting BEARISH trades with strike_price != 0 for both Afternoon and Morning
        option_entry = next((float(trade['avg_prc']) for trade in afternoon_trades if trade['direction'] == 'BEARISH' and trade['strike_price'] != "0"), None)
        option_exit = next((float(trade['avg_prc']) for trade in morning_trades if trade['direction'] == 'BEARISH' and trade['strike_price'] != "0"), None)
    
    
    if broker == "zerodha":
        future_tax = zerodha_futures_taxes(qty, future_entry, future_exit, 1)
        option_tax = zerodha_taxes(qty, option_entry, option_exit, 1)
        total_tax = future_tax + option_tax
    elif broker == "aliceblue":
        future_tax = aliceblue_futures_taxes(qty, future_entry, future_exit, 1)
        option_tax = aliceblue_taxes(qty, option_entry, option_exit, 1)
        total_tax = future_tax + option_tax

    # Calculating trade points based on direction
    direction = afternoon_trades[0]["direction"]
    if direction == "BULLISH":
        trade_points = (future_exit - future_entry) + (option_exit - option_entry)
    elif direction == "BEARISH":  # Assuming BEARISH
        trade_points = (future_entry - future_exit) + (option_exit - option_entry)
    PnL = trade_points * qty

    # Appending to result list
    trade_data = {
        "Trade_Type": direction,
        "Qty": qty,
        "Future_Entry": future_entry,
        "Future_Exit": future_exit,
        "Option_Entry": option_entry,
        "Option_Exit": option_exit,
        "Trade_Points": trade_points,
        "PnL": PnL,
        "Tax": total_tax,
    }
    result.append(trade_data)
    return result

script_dir = os.path.dirname(os.path.realpath(__file__))
broker_filepath = os.path.join(script_dir, "broker.json")
json_dir = os.path.join(script_dir, '..','UserProfile',"json")
excel_dir = os.path.join(script_dir, '..','UserProfile',"excel")

with open(broker_filepath) as file:
    data = json.load(file)

# Initialize an empty list for the accounts to trade
user_list = []

# Go through each broker
for broker, broker_data in data.items():
    # Check if 'accounts_to_trade' is in the broker data
    if 'accounts_to_trade' in broker_data:
        # Add each account to the list
        for account in broker_data['accounts_to_trade']:
            user_list.append((broker, account))

for broker, user in user_list:
    # Load the JSON data
    with open(os.path.join(json_dir, f"{user}.json")) as file:
        user_data = json.load(file)

    phone_number = data[broker][user]["mobile_number"]

    # Default values
    amipy_df = pd.DataFrame()
    overnight_df = pd.DataFrame()
    amipy_pnl = 0
    amipy_tax = 0
    overnight_options_pnl = 0
    overnight_options_tax = 0

    if "MPWizard" in user_data[broker]["orders"]:
        # Process the MPWizard trades
        mpwizard_data = process_mpwizard_trades(user_data[broker]["orders"]["MPWizard"])
        mpwizard_df = pd.DataFrame(mpwizard_data)
        mpwizard_pnl = round(mpwizard_df["PnL"].sum(),2)
        mpwizard_tax = round(mpwizard_df["Tax"].sum(),2)

    if "Amipy" in user_data[broker]["orders"]:
        amipy_data_short = []
        amipy_data_long = []
        if "ShortSignal" in user_data[broker]["orders"]["Amipy"]:
            # Process the AmiPy ShortSignal and ShortCoverSignal trades
            amipy_data_short = process_short_trades(user_data[broker]["orders"]["Amipy"]["ShortSignal"], 
                                                    user_data[broker]["orders"]["Amipy"]["ShortCoverSignal"])
        if "LongSignal" in user_data[broker]["orders"]["Amipy"]:
            # Process the AmiPy LongSignal and LongCoverSignal trades
            amipy_data_long = process_long_trades(user_data[broker]["orders"]["Amipy"]["LongSignal"], 
                                                user_data[broker]["orders"]["Amipy"]["LongCoverSignal"])

        # Combine short and long trades into a single DataFrame
        amipy_data = amipy_data_short + amipy_data_long
        if amipy_data:
            amipy_df = pd.DataFrame(amipy_data)
            amipy_pnl = round(amipy_df["PnL"].sum(), 2)
            amipy_tax = round(amipy_df["Tax"].sum(), 2)
    
    if "Overnight_Options" in user_data[broker]["orders"]:
        # Process the Overnight_Options trades
        overnight_options_data = process_overnight_options_trades(user_data[broker]["orders"]["Overnight_Options"])
        if overnight_options_data:
            overnight_options_df = pd.DataFrame(overnight_options_data)
            overnight_options_pnl = round(overnight_options_df["PnL"].sum(),2)
            overnight_options_tax = round(overnight_options_df["Tax"].sum(),2)

    # # Read existing data from Excel file into separate DataFrames
    mpwizard_existing_df = pd.read_excel(os.path.join(excel_dir, f"{user}.xlsx"), sheet_name="MPWizard")
    amipy_existing_df = pd.read_excel(os.path.join(excel_dir, f"{user}.xlsx"), sheet_name="AmiPy")
    overnight_existing_df = pd.read_excel(os.path.join(excel_dir, f"{user}.xlsx"), sheet_name="Overnight_options")

    # # Append new data
    mpwizard_final_df = pd.concat([mpwizard_existing_df, mpwizard_df])
    amipy_final_df = pd.concat([amipy_existing_df, amipy_df])
    overnight_final_df = pd.concat([overnight_existing_df, overnight_options_df])
    # overnight_final_df = pd.concat([overnight_existing_df])

    gross_pnl = mpwizard_pnl + amipy_pnl + overnight_options_pnl
    # gross_pnl = mpwizard_pnl + amipy_pnl 
    tax = mpwizard_tax + amipy_tax + overnight_options_tax
    # tax = mpwizard_tax + amipy_tax
    net_pnl = gross_pnl - tax

    current_capital = data[broker][user]['current_capital']
    if net_pnl > 0:
        expected_capital = current_capital + net_pnl
    else:
        expected_capital = current_capital - abs(net_pnl)

    message_parts = [f"Hello {user},We hope you're enjoying a wonderful day.\n Here are your PNLs for today:\n"]

    if "MPWizard" in user_data[broker]["orders"]:
        message_parts.append(f"MPWizard: {custom_format(mpwizard_pnl)}")

    if "Amipy" in user_data[broker]["orders"]:
        message_parts.append(f"AmiPy: {custom_format(amipy_pnl)}")

    if "Overnight_Options" in user_data[broker]["orders"]:
        message_parts.append(f"Overnight Options: {custom_format(overnight_options_pnl)}")

    message_parts.append(f"\n**Gross PnL: {custom_format(gross_pnl)}**")
    message_parts.append(f"**Expected Tax: {custom_format(tax)}**")
    message_parts.append(f"**Current Capital: {custom_format(current_capital)}**")
    message_parts.append(
        f"**Expected Morning Balance : {custom_format(expected_capital)}**")
    message_parts.append("\nBest Regards,\nSerendipity Trading Firm")

    message = "\n".join(message_parts)
    message = message.replace('\u20b9', 'INR')


    message = "\n".join(message_parts)
    print(message)

    # Save data to broker.json
    data_to_store = {
        'Total PnL': net_pnl,
        'Current Capital': current_capital,
        'Expected Morning Balance': expected_capital
    }
    user_details = data[broker][user]
    user_details["yesterday_PnL"] = net_pnl
    user_details["expected_morning_balance"] = round(expected_capital,2)
    data[broker][user] = user_details


    with open(broker_filepath, 'w') as json_file:
        json.dump(data, json_file, indent=4)

    # send discord message
    script_dir = os.path.dirname(os.path.abspath(__file__))   
    parent_dir = os.path.abspath(os.path.join(script_dir, '..','..'))
    filepath = os.path.join(parent_dir, '+918618221715.session')
    
    with TelegramClient(filepath, api_id, api_hash) as client:
        client.send_message(phone_number, message, parse_mode='md')

    # Load existing workbook
    excel_path = os.path.join(excel_dir, f"{user}.xlsx")
    book = load_workbook(excel_path)

    # Read existing sheets into DataFrames, except for the ones we want to replace
    existing_dfs = {}
    for sheet_name in book.sheetnames:
        if sheet_name not in ['MPWizard', 'AmiPy', 'Overnight_options']:
            existing_dfs[sheet_name] = pd.read_excel(excel_path, sheet_name=sheet_name)

    # Create a temporary new Excel file
    temp_path = os.path.join(excel_dir, f"{user}_new.xlsx")
    with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
        # Write existing sheets
        for sheet_name, df in existing_dfs.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Write new sheets
        mpwizard_final_df.to_excel(writer, sheet_name='MPWizard', index=False)
        amipy_final_df.to_excel(writer, sheet_name='AmiPy', index=False)
        overnight_final_df.to_excel(writer, sheet_name='Overnight_options', index=False)

    # Delete the old file and rename the new one
    os.remove(excel_path)
    os.rename(temp_path, excel_path)
