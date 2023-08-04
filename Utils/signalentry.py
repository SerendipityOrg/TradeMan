###### Working DONT NOT TOUCH ######


import json
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import os


script_dir = os.path.dirname(os.path.realpath(__file__))
signal_excel = os.path.join(script_dir, "excel","signal.xlsx")


# Function to load JSON data
def load_json(filepath):
    with open(filepath, 'r') as json_file:
        data = json.load(json_file)
    return data

# Function to handle MPWizard data
def write_to_excel_MPWizard(data, sheet_name, workbook):
    # Get or create sheet
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

    # Find the last filled row
    last_row = sheet.max_row

    for index, entry in enumerate(data['indices'], start=1):
        # Extract information
        trade_no = last_row + index
        option = entry['SignalEntry']['Option']
        event = entry['SignalEntry']['Event']
        entry_time = entry['SignalEntry']['EntryTime']
        exit_time = entry['SignalEntry']['ExitTime']
        entry_price = entry['SignalEntry']['EntryPrice']
        exit_price = entry['SignalEntry']['ExitPrice']
        trade_points = entry_price - exit_price

        # Append to Excel
        row = [trade_no, option, event, entry_time, exit_time, entry_price, exit_price, trade_points]
        sheet.append(row)

        # Color cell based on trade points
        if trade_points > 0:
            fill = PatternFill(start_color="00FF00", fill_type="solid")  # Green
        else:
            fill = PatternFill(start_color="FF0000", fill_type="solid")  # Red

        sheet[get_column_letter(8) + str(sheet.max_row)].fill = fill

# Function to handle AmiPy data
def write_to_excel_AmiPy(data, sheet_name, workbook):
    # Get or create sheet
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

    # Find the last filled row
    last_row = sheet.max_row

    for index, entry in enumerate(data['Nifty'], start=1):
        # Extract information
        trade_no = last_row + index
        strategy = "AmiPy"
        trade_type = None
        strike_price = None
        entry_time = None
        exit_time = None
        entry_price = None
        exit_price = None
        trade_points = None
        date = None

        signal_entry = entry['SignalEntry']
        if 'ShortSignal' in signal_entry:
            trade_type = "ShortSignal"
            date = signal_entry['ShortSignal']['Date']
            strike_price = signal_entry['ShortSignal']['Strike_Price']
            entry_time = signal_entry['ShortSignal']['TradeEntryTime']
            entry_price = signal_entry['ShortSignal']['TradeEntryPrice']
        if 'ShortCoverSignal' in signal_entry:
            exit_time = signal_entry['ShortCoverSignal']['TradeExitTime']
            exit_price = signal_entry['ShortCoverSignal']['TradeExitPrice']
            trade_points = entry_price - exit_price  # For ShortSignal
        elif 'LongSignal' in signal_entry:
            trade_type = "LongSignal"
            date = signal_entry['LongSignal']['Date']
            strike_price = signal_entry['LongSignal']['Strike_Price']
            entry_time = signal_entry['LongSignal']['TradeEntryTime']
            entry_price = signal_entry['LongSignal']['TradeEntryPrice']
        if 'LongCoverSignal' in signal_entry:
            exit_time = signal_entry['LongCoverSignal']['TradeExitTime']
            exit_price = signal_entry['LongCoverSignal']['TradeExitPrice']
            trade_points = exit_price - entry_price  # For LongSignal

        # Append to Excel
        row = [trade_no, strategy, trade_type, date, strike_price, entry_time, exit_time, entry_price, exit_price, trade_points]
        sheet.append(row)

        # Color cell based on trade points
        if trade_points > 0:
            fill = PatternFill(start_color="00FF00", fill_type="solid")  # Green
        else:
            fill = PatternFill(start_color="FF0000", fill_type="solid")  # Red

        sheet[get_column_letter(10) + str(sheet.max_row)].fill = fill

# Load workbook
# wb = load_workbook('example.xlsx')

# Check if the file exists
if not os.path.isfile(signal_excel):
    # If the file doesn't exist, create a new one
    wb = Workbook()
    wb.save(signal_excel)

# Now you can safely load the workbook
wb = load_workbook(signal_excel)

script_dir = os.path.dirname(os.path.abspath(__file__))
MPWizard_filepath = os.path.join(script_dir, '..', 'MPWizard', 'MPWizard.json')
Amipy_filepath = os.path.join(script_dir, '..', 'Amipy','AmiPy.json')

# Load data and write to excel
data_MPWizard = load_json(MPWizard_filepath)
write_to_excel_MPWizard(data_MPWizard, 'MPWizard', wb)

data_AmiPy = load_json(Amipy_filepath)
write_to_excel_AmiPy(data_AmiPy, 'AmiPy', wb)

# Save workbook
wb.save(signal_excel)
